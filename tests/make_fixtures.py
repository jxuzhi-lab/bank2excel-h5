# -*- coding: utf-8 -*-
"""make_fixtures.py —— M2 构建期脚本: 基准 xlsx → fixtures JSON。

用法:
    python make_fixtures.py            # 在本地 CPython 环境跑一次
产出:
    tests/fixtures/manifest.json       样本清单(pdf 相对路径 + fixture 文件名)
    tests/fixtures/sample_NN.json      每样本的规范化单元格矩阵
说明:
    单元格值经 norm() 规范化(datetime→文本 与基准旧格式兼容、数字→str(float)),
    与 m1_check.py 的对比语义完全一致; 浏览器端 regression.html 用同一套 norm
    对 Pyodide 转换结果逐格比对。
    本地 fixtures 含真实数据, 仅限本机测试; 公开部署前须按 M6 脱敏重生成。
"""
import glob
import json
import os
import sys
from datetime import date, datetime

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))          # bank2excel-h5/tests
H5_DIR = os.path.dirname(HERE)                             # bank2excel-h5
ROOT = os.path.dirname(H5_DIR)                             # 工作区根(含 测试样本/)
SAMPLES_DIR = os.path.join(ROOT, "测试样本")
FIXTURES_DIR = os.path.join(HERE, "fixtures")


def norm(v):
    """规范化单元格值 → 可 JSON 化的字符串(m1_check.py 同款语义)。"""
    if isinstance(v, datetime):
        return v.strftime("%Y%m%d %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%Y%m%d")
    if v is None:
        return ""
    try:
        return str(float(v))
    except Exception:
        return str(v).strip()


def main():
    os.makedirs(FIXTURES_DIR, exist_ok=True)
    pdfs = sorted(glob.glob(os.path.join(SAMPLES_DIR, "*.pdf")))
    if not pdfs:
        print("未找到测试样本:", SAMPLES_DIR)
        return 2
    manifest = []
    idx = 0
    for p in pdfs:
        base = os.path.splitext(p)[0] + ".xlsx"
        if not os.path.exists(base):
            print("[跳过] 无基准 xlsx:", os.path.basename(p))
            continue
        idx += 1
        ws = openpyxl.load_workbook(base).active
        cells = []
        for row in ws.iter_rows():
            cells.append([norm(c.value) for c in row])
        slug = f"sample_{idx:02d}.json"
        with open(os.path.join(FIXTURES_DIR, slug), "w", encoding="utf-8") as f:
            json.dump({"rows": ws.max_row, "cols": ws.max_column,
                       "cells": cells}, f, ensure_ascii=False,
                      separators=(",", ":"))
        manifest.append({
            "name": os.path.basename(p),
            "pdf": "pdf/" + os.path.basename(p),
            "fixture": "fixtures/" + slug,
            "rows": ws.max_row,
        })
        print(f"[OK] {slug}  <- {os.path.basename(p)}  ({ws.max_row} 行)")
    with open(os.path.join(FIXTURES_DIR, "manifest.json"), "w",
              encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=1)
    print(f"\n共 {len(manifest)} 个 fixture -> {FIXTURES_DIR}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
