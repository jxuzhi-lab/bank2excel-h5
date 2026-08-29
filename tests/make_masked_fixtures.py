# -*- coding: utf-8 -*-
"""make_masked_fixtures.py —— M6 脱敏回归集生成 + 验证。

流程:
  1. 用脱敏 PDF (tests/pdf_masked) 跑 shim.convert_bytes
  2. 提取结果既当"转换输出"又当"基准"(自洽): 只验证
     a) 脱敏 PDF 仍可被管道正常提取(表头/记录/日期形态健壮性)
     b) 输出无泄漏(≥8 位连续数字 = 0)
  3. 生成 tests/fixtures_masked/ (脱敏基准 JSON + manifest)

用法: python make_masked_fixtures.py
"""
import glob
import io
import json
import os
import sys
from datetime import date, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(HERE), "python"))
os.environ["PYODIDE"] = "1"

import shim  # noqa: E402
import extract_bank_statement as eng  # noqa: E402

eng.PYODIDE = True
import openpyxl  # noqa: E402
import pymupdf  # noqa: E402

DST = os.path.join(HERE, "pdf_masked")
FXT = os.path.join(HERE, "fixtures_masked")


def norm(v):
    if isinstance(v, datetime):
        return v.strftime("%Y%m%d %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%Y%m%d")
    if v is None:
        return ""
    try:
        return str(float(v))
    except Exception:
        return str(v).strip()


def leak_scan(pdf_path):
    """深层泄漏扫描: 提取全文文本找 ≥6 位连续数字(账号/流水号残漏)。"""
    doc = pymupdf.open(pdf_path)
    bad = []
    for pno, page in enumerate(doc, 1):
        for w in page.get_text("words"):
            for token in w[4].split():
                if len(token) >= 6 and token.isdigit():
                    bad.append((pno, token))
    doc.close()
    return bad


def main():
    os.makedirs(FXT, exist_ok=True)
    pdfs = sorted(glob.glob(os.path.join(DST, "*.pdf")))
    manifest, ok = [], 0
    for i, p in enumerate(pdfs, 1):
        name = os.path.basename(p)
        try:
            with open(p, "rb") as f:
                r = shim.convert_bytes(f.read())
            wb = openpyxl.load_workbook(io.BytesIO(r["xlsx"]))
            ws = wb.active
            cells = [[norm(c.value) for c in row] for row in ws.iter_rows()]
            slug = f"sample_{i:02d}.json"
            with open(os.path.join(FXT, slug), "w", encoding="utf-8") as f:
                json.dump({"rows": ws.max_row, "cols": ws.max_column,
                           "cells": cells}, f, ensure_ascii=False,
                          separators=(",", ":"))
            manifest.append({"name": name, "pdf": "pdf/" + name,
                             "fixture": "fixtures/" + slug, "rows": ws.max_row})
            bad = leak_scan(p)
            status = "OK " if not bad else "LEAK"
            ok += (not bad)
            extra = f" leak={bad[:2]}" if bad else ""
            print(f"{status} {name}  {r['rows']} rows{extra}")
        except Exception as e:  # noqa: BLE001
            print(f"FAIL {name}: {str(e)[:90]}")
    with open(os.path.join(FXT, "manifest.json"), "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=1)
    print(f"\n{ok}/{len(pdfs)} clean -> {FXT}")
    return 0 if ok == len(pdfs) else 1


if __name__ == "__main__":
    sys.exit(main())
