# -*- coding: utf-8 -*-
"""m1_check.py —— M1 回归验证(本地 CPython 冒烟)。

用法:
    python m1_check.py                 # 跑全部 13 样本 vs 基准 xlsx
验收标准:
    13/13 PASS(其中 22建行/建行-1-7月 基准为旧版文本日期格式,
    用 norm() 规范化 datetime→文本 后逐格 0 差异)。
说明:
    本测试在本地 CPython 跑 PYODIDE=1 分支(与浏览器内引擎代码路径一致),
    验证 shim.convert_bytes 的字节流进出改造未破坏提取逻辑。
    浏览器端回归(M2)复用同一套逻辑, 换成 pyodide 运行时。
"""
import glob
import os
import sys
from datetime import date, datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ["PYODIDE"] = "1"

import shim  # noqa: E402
import extract_bank_statement as eng  # noqa: E402

eng.PYODIDE = True

import openpyxl  # noqa: E402

SAMPLES_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "测试样本")


def norm(v):
    """规范化对比: datetime → 基准的文本格式(旧版基准兼容), 其余转字符串。"""
    if isinstance(v, datetime):
        return v.strftime("%Y%m%d %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%Y%m%d")
    if v is None:
        return ""
    try:
        return str(float(v))
    except Exception:
        return str(v).strip()


def cell_eq(a, b):
    return norm(a) == norm(b)


def main():
    pdfs = sorted(glob.glob(os.path.join(SAMPLES_DIR, "*.pdf")))
    if not pdfs:
        print("未找到测试样本:", SAMPLES_DIR)
        return 2
    results, npass = [], 0
    for p in pdfs:
        name = os.path.basename(p)
        base = os.path.splitext(p)[0] + ".xlsx"
        if not os.path.exists(base):
            results.append((name, "NO_BASELINE", 0))
            continue
        try:
            with open(p, "rb") as f:
                r = shim.convert_bytes(f.read())
            tmp = "/tmp/m1_check.xlsx"
            with open(tmp, "wb") as f:
                f.write(r["xlsx"])
            ws1 = openpyxl.load_workbook(tmp).active
            ws2 = openpyxl.load_workbook(base).active
            diff = 0
            for row in range(1, max(ws1.max_row, ws2.max_row) + 1):
                for col in range(1, max(ws1.max_column, ws2.max_column) + 1):
                    if not cell_eq(ws1.cell(row, col).value,
                                   ws2.cell(row, col).value):
                        diff += 1
            tag = "PASS" if diff == 0 else f"DIFF({diff})"
            if diff == 0:
                npass += 1
            results.append((name, tag, r["rows"]))
        except Exception as e:  # noqa: BLE001
            results.append((name, f"ERROR: {str(e)[:60]}", 0))

    print(f"\n{'样本':<44}{'结果':<22}{'行数':>6}")
    print("-" * 74)
    for name, tag, rows in results:
        print(f"{name:<44}{tag:<22}{rows:>6}")
    print("-" * 74)
    print(f"通过 {npass}/{len(results)}")
    return 0 if npass == len(results) else 1


if __name__ == "__main__":
    sys.exit(main())
