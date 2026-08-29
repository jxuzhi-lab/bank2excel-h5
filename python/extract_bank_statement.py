#!/usr/bin/env python3
"""
银行对账单 PDF → Excel 转换脚本
================================
自动识别表头行、列边界、记录边界，忽略水印/印章（基于文字层提取），
将多页对账单合并为一张 Excel 工作表，并打印提取统计（笔数、借贷方/收支合计）。
交付前由代理按 SKILL.md「交付前抽样比对」人工核对 PDF 拼接处（跨页首末条、
粘合日期、页脚相邻行等）；不再内置余额连续性等自动校验（2026-08-18 起）。

依赖: pymupdf, openpyxl
用法:
  python extract_bank_statement.py --input statement.pdf [--output out.xlsx] [选项]

选项:
  --input <pdf>            必填. 输入对账单 PDF 路径
  --output <xlsx>          可选. 输出路径(默认与 PDF 同目录同名 .xlsx)
  --date-pattern <regex>   可选. 记录起始日期行匹配正则(默认自动探测 4 种)
  --header-keywords <k1,k2> 可选. 表头识别关键词, 逗号分隔
  --password <pw>          可选. PDF 打开密码(加密对账单, 如华夏银行)
  --sheet <name>           可选. 工作表名(默认取 PDF 文件名)
  --no-auto-onboard        可选. 关闭失败时自动 onboarding(2026-08-20 新增)
  --no-diag                可选. 关闭失败诊断包 .diag.json(2026-08-20 新增)
  --keep-text              可选. 日期/金额列保留 PDF 原文文本(默认输出为 Excel 日期/数值格式)
  --debug                  可选. 输出诊断信息

退出码: 0=成功, 2=处理出错(提取失败)。数据质量靠交付前人工抽样比对把关。

性能说明(2026-08-17 优化):
  - PDF 只打开一次(extract_statement 接收已打开的 doc, 不再二次 open)
  - 每页记录切分用排序 + bisect 二分(原为每条记录全量扫描 region)
  - 每页 find_date_rows 只跑一遍, region 内按 y 过滤复用
  - 统计直接作用于内存 ws, 不再保存后重新 load_workbook
  - META/SUMMARY 关键词预编译为正则
"""
import argparse
import bisect
import hashlib
import io
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime

import pymupdf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.cell_style import StyleArray
from openpyxl.utils import get_column_letter

# ---------- 运行环境分支(2026-08-28 H5 改造) ----------
# PYODIDE=1 时: 字节流进出(open_pdf 走 stream)、xlsx 写入内存、无诊断落盘、
#              无视觉兜底(运行时无 LLM/外部视觉, 纯规则路径)、日志收集进内存列表。
PYODIDE = os.environ.get("PYODIDE", "") == "1"
LOG_BUFFER = []  # PYODIDE 模式下 log() 输出收集于此, 随转换结果返回前端
PROGRESS_CB = None  # 页级进度回调 fn(current_page, total_pages); H5 由 shim 注入


def set_progress_cb(cb):
    """注册/清除进度回调(H5 用): set_progress_cb(None) 复位。"""
    global PROGRESS_CB
    PROGRESS_CB = cb


def open_pdf(src, password=None):
    """统一的 PDF 打开入口: 本地路径走 pymupdf.open(path);
    Pyodide/字节流场景传 bytes/bytearray, 走 Document(stream=...)。
    password 仅本地路径场景需要(stream 场景在调用处已处理认证)。"""
    if isinstance(src, (bytes, bytearray)):
        doc = pymupdf.Document(stream=bytes(src))
        if doc.needs_pass and password:
            if not doc.authenticate(password):
                raise RuntimeError("PDF 密码错误")
        return doc
    return pymupdf.open(src)

# ---------- 默认配置 ----------
DEFAULT_HEADER_KEYWORDS = [
    "交易时间", "时间", "日期", "摘要", "用途", "备注",
    "凭证类型", "凭证号码", "凭证号",
    "借方发生额", "借方金额", "支出", "转出",
    "贷方发生额", "贷方金额", "收入", "转入",
    "账户余额", "余额", "流水号",
    "对方户名/账号", "对方户名", "对方账号", "对方行名", "开户行",
]
# 表头合并词拆分时使用的标准列名（按顺序匹配）
# 2026-08-17 扩: 建行活期明细 17 列(账号/币种/对方开户机构/记账日期/凭证种类/凭证号/交易介质编号等)
# 2026-08-18 扩: 北京农商银行(标志交易金额/对方银行名称/交易渠道/交易机构/现转标志/附言)
# 2026-08-21 扩: 浦发银行两级表头(父表头"发生额"横跨子列"借方/贷方", 子列名单独成词;
#   "借方/贷方"入 SPLIT_NAMES 同时让 hit_max 延伸到子列名行, 表头带才能含子列词)
SPLIT_NAMES = [
    "交易时间戳", "交易时间", "借方发生额", "贷方发生额", "借方", "贷方", "账户余额", "余额",
    "标志交易金额", "交易金额",
    "对方户名", "对方账号", "对方行名", "开户行", "对方开户机构", "对方银行名称",
    "交易渠道", "交易机构", "现转标志",
    "凭证号码", "凭证种类", "凭证类型", "凭证号",
    "流水号", "时间", "日期", "记账日期", "摘要", "备注",
    "账号", "币种", "交易介质编号", "附言",
]
DEFAULT_DATE_PATTERNS = [
    re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$"),
    re.compile(r"^\d{4}-\d{1,2}/\d{1,2}$"),
    re.compile(r"^\d{4}-\d{1,2}-\d{1,2}$"),
    # 8 位纯数字日期必须形如 19xx/20xx 且月日合法: 建行企业账户明细(17 列)把
    # 本方账号/对方账号/凭证号等也拆成 8 位数字词(11050172/22010274/10732590),
    # 纯 ^\d{8}$ 会把每条记录在多个 8 位数字处切成碎片(记录数虚高、字段跨记录错乱)。
    # 真实日期(交易时间/记账日期)全部为 YYYYMMDD 合法日历日期, 收紧不误伤。
    re.compile(r"^(?:19|20)\d{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01])$"),
]
# 摘要+日期粘合词(2026-08-18 建行"全部交易明细"变体): PDF 把交易日期拼进摘要词尾,
# 如 "支付机构提现20250103"/"银联无卡支付20260811", 无独立日期词会导致该条记录
# 没有日期锚点、被并进上一条记录。CJK 前缀限定保证不误伤纯账号数字
# (如 "2088512815070261/杜晓武" 尾部是户名而非日期); 页眉 "起止日期:...-20260818"
# 同样命中但 y 在数据区上方, 会被 region 过滤排除。
GLUED_DATE_RE = re.compile(r"^(.*[\u4e00-\u9fff])((?:19|20)\d{6})$")
# 脱敏占位词(纯星号, 如建行"交易地点/附言"列的 "***"): 是银行对数据的脱敏显示,
# 不是水印/页脚。跨页同位置出现时不可按固定元素删除。
MASK_WORD_RE = re.compile(r"^\*+$")
# 对角印章/防伪码(ICBC 借记明细 "FFFEB5EC2026"/"0E69757E9026"/"1AD4AD6F1026" 等):
# 12 位大写十六进制且含字母(账号/摘要码/序号均为纯数字, 不会误匹配)。与
# "非数据行 y" 联合判定(见 extract_statement)后剔除; 印章防伪码每页不同,
# 不满足跨页固定元素阈值, 需单独按形态识别。
STAMP_CODE_RE = re.compile(r"^(?=.*[A-F])[0-9A-F]{10,}$")
# 2026-08-22 补充: 工行还有 12 位**纯数字**对角防伪码(如 892463878026),
# 每页只出现 1 次甚至只有最后一页出现, 不能靠跨页固定阈值识别。
# 该类词由大字旋转印章渲染, 归一化后字高通常 >15pt; 常规 12 位账号/流水
# 单元格字高约 5-8pt, 因此用"纯 12 位数字 + 异常高字盒"联合判定。
PURE_NUMERIC_STAMP_RE = re.compile(r"^\d{12}$")
STAMP_WORD_H = 15.0


def _is_stamp_code_word(w):
    """判断文字层词是否为对角印章/防伪码词。

    十六进制防伪码维持原判定; 纯 12 位数字必须同时是异常大字号/大高度词,
    避免把正常 12 位账号、交易流水号或日期编号误删。
    """
    text = str(w[4] or "").strip()
    if STAMP_CODE_RE.match(text):
        return True
    if PURE_NUMERIC_STAMP_RE.match(text):
        return (float(w[3]) - float(w[1])) > STAMP_WORD_H
    return False
AMOUNT_RE = re.compile(r"^-?\d{1,3}(,\d{3})*\.\d{2}$")
# 金额列内的时间形态词(民生个人版对账单: 摘要嵌入时间被 PDF 拆词后 x 落入金额列,
# 如 '银联入账-自动提现+2026-01-03' 与 '05:54:02' 拆开, 时间词 x=244 进"交易金额"列,
# 拼出 '0554023,817.46' 式脏值)。形态: HH:MM:SS 或 HHMMSS(6 位时分秒合法)。
TIME_WORD_RE = re.compile(r"^\d{2}:\d{2}:\d{2}$")
COMPACT_TIME_RE = re.compile(r"^\d{6}$")


def is_time_word(text):
    """判断词是否为纯时间形态(HH:MM:SS 或 6 位紧凑 HHMMSS 且时分秒合法)。
    金额列出现的这类词几乎必然是 PDF 拆词, 应重分类到摘要列。"""
    if TIME_WORD_RE.match(text):
        return True
    if COMPACT_TIME_RE.match(text):
        h, m, s = int(text[0:2]), int(text[2:4]), int(text[4:6])
        return h <= 23 and m <= 59 and s <= 59
    return False

# 对手机构代码词形态（村镇银行/城商行流水常见，如 %1000050201%02%99%%000）
# 完整形态 / PDF 显示截断形态（%1...）
COUNTERPART_CODE_RE = re.compile(r"^%\d[\d%]*%+\d*$")
TRUNCATED_CODE_RE = re.compile(r"^%\d.*\.\.\.$")
# 页脚页码形态（"第 3 页"/"共 131 页"/"第28/28页"/整行"18/18"）。
# 用正则避免"共"字误伤数据（如"北京公共交通控股…"含"共"字）。
# 注意: 纯数字斜杠必须整行锚定, 否则会误匹配日期"2025/01/03"(民生银行日期格式)。
PAGE_FOOTER_RE = re.compile(
    r"第\s*\d+\s*(/\s*\d+)?\s*页|共\s*\d+\s*页|^\s*\d{1,3}\s*/\s*\d{1,3}\s*$"
)
# 页眉/页脚行关键词（含这些词的行不进入数据区）。预编译为正则, 等价于原 any(k in text) 子串匹配
META_KEYWORDS = [
    "打印渠道", "打印柜员", "打印时间", "打印日期", "生成时间",
    "客户名称", "客户账号", "币种", "开户机构", "起止日期", "对账单",
    "温馨提示", "具体交易详情", "说明：", "查询时间", "查询完毕",
    # 分页汇总/回单页脚(旅立方等): "本页汇总：统计：…"、"本页回单：第1-15笔 共34笔"
    "本页汇总", "本页回单",
    # 招商银行末页"合并统计"块(2026-08-19): "合并统计/合并收入(+)/合并支出(-)"
    # 在末条记录下方, 不排除会被吸进末条记录的 货币/金额/摘要 列。
    "合并统计", "合并收入", "合并支出",
]
META_RE = re.compile("|".join(re.escape(k) for k in META_KEYWORDS), re.I)

# 元信息行(非交易, 整行排除): 查询时间/打印日期/分页汇总统计等
META_ROW_RE = re.compile(
    r"查询时间|打印日期"
    r"|本页汇总|本页回单|^统计：",
    re.I,
)


# 页脚装饰横线(10+ 连续横线/下划线字符, 如招商银行末页整行横线
# '————————————…' 53 个 U+2500)。数据行不会出现 10+ 连续横线, 安全。
# 2026-08-17 扩 '='(北京农商银行末页整行 ===)。
FOOTER_LINE_RE = re.compile(r"[─━—―\-_=]{10,}")


def is_date_like_word(text):
    """判断词是否为完整日期形态(YYYYMMDD / YYYY-MM-DD / YYYY/MM/DD 等)。

    用途: 固定元素(水印)判定保护。日期列词必须原样保留 —— 即使它跨页出现在
    相同坐标(每页首条记录布局一致的长文档很常见), 也绝不能当"水印"删除。
    (浦发 144 页流水: 每页首条交易日期 '20240101' y=46.3, 若不加保护会被
    clean_overlaps 整列清空 → 首条记录日期丢失。)"""
    if not text:
        return False
    t = text.strip()
    if re.fullmatch(r"(?:19|20)\d{2}[./-]?\d{1,2}[./-]?\d{1,2}", t)             and not is_time_word(t):
        return True
    return False

# 对手机构代码词形态（村镇银行/城商行流水常见，如 %1000050201%02%99%%000）
# 完整形态 / PDF 显示截断形态（%1...）
COUNTERPART_CODE_RE = re.compile(r"^%\d[\d%]*%+\d*$")
TRUNCATED_CODE_RE = re.compile(r"^%\d.*\.\.\.$")
# 页脚页码形态（"第 3 页"/"共 131 页"/"第28/28页"/整行"18/18"）。
# 用正则避免"共"字误伤数据（如"北京公共交通控股…"含"共"字）。
# 注意: 纯数字斜杠必须整行锚定, 否则会误匹配日期"2025/01/03"(民生银行日期格式)。
PAGE_FOOTER_RE = re.compile(
    r"第\s*\d+\s*(/\s*\d+)?\s*页|共\s*\d+\s*页|^\s*\d{1,3}\s*/\s*\d{1,3}\s*$"
)
# 页眉/页脚行关键词（含这些词的行不进入数据区）。预编译为正则, 等价于原 any(k in text) 子串匹配
META_KEYWORDS = [
    "打印渠道", "打印柜员", "打印时间", "打印日期", "生成时间",
    "客户名称", "客户账号", "币种", "开户机构", "起止日期", "对账单",
    "温馨提示", "具体交易详情", "说明：", "查询时间", "查询完毕",
    # 分页汇总/回单页脚(旅立方等): "本页汇总：统计：…"、"本页回单：第1-15笔 共34笔"
    "本页汇总", "本页回单",
    # 招商银行末页"合并统计"块(2026-08-19): "合并统计/合并收入(+)/合并支出(-)"
    # 在末条记录下方, 不排除会被吸进末条记录的 货币/金额/摘要 列。
    "合并统计", "合并收入", "合并支出",
]
META_RE = re.compile("|".join(re.escape(k) for k in META_KEYWORDS), re.I)

# 元信息行(非交易, 整行排除): 查询时间/打印日期/分页汇总统计等
META_ROW_RE = re.compile(
    r"查询时间|打印日期"
    r"|本页汇总|本页回单|^统计：",
    re.I,
)


# 页脚装饰横线(10+ 连续横线/下划线字符, 如招商银行末页整行横线
# '————————————…' 53 个 U+2500)。数据行不会出现 10+ 连续横线, 安全。
# 2026-08-17 扩 '='(北京农商银行末页整行 ===)。
FOOTER_LINE_RE = re.compile(r"[─━—―\-_=]{10,}")


def is_footer_word(text):
    """判断是否为页脚/页眉词：命中页码形态正则、装饰横线 或 META_KEYWORDS。
    2026-08-19 修正: 装饰横线须**横线字符占比 ≥90% 且不含日期词**才算页脚
    (招商末页 53 个 U+2500 / 北京农商 整行 === / KA0200… 页尾 "----END----" 93%);
    数据行即使空单元格多为横线占位(短信费末条 78%)也**不是页脚**——否则
    每页末条记录被切掉(1334 → 1285 笔); 而数据行内嵌的 '-' 占位符占比 ~25%,
    早前版本因 ≥60% 规则把每页第 5 条后的行当页脚切掉(440 笔)。"""
    if PAGE_FOOTER_RE.search(text) or META_RE.search(text):
        return True
    if FOOTER_LINE_RE.search(text):
        stripped = re.sub(r"\s", "", text)
        if not stripped:
            return False
        dash_n = sum(1 for ch in stripped if ch in "─━—―-_=")
        if dash_n / len(stripped) < 0.9:
            return False
        # 含日期词的行是数据行(如末条 "2026-08-08 03:45:35 …")不是页脚
        return not any(p.match(t) for t in re.split(r"[\s,，:：]+", text)
                       for p in DEFAULT_DATE_PATTERNS)
    return False


# 汇总行关键词
SUMMARY_KEYWORDS = ["借方发生总额", "贷方发生总额", "合计笔数", "期末余额", "期初余额",
                    "借方累计", "贷方累计", "总笔数",
                    "总收入笔数", "总收入金额", "总支出笔数", "总支出金额"]
SUMMARY_RE = re.compile("|".join(re.escape(k) for k in SUMMARY_KEYWORDS), re.I)
# 余额类汇总行(期初/期末余额): 即使带日期也整行排除
BAL_SUMMARY_RE = re.compile(
    r"期初余额|期末余额",
    re.I,
)


def normalize_page_words(page, words):
    """把 mediabox 坐标系(未旋转)的 words 转换为视觉坐标系(rect 坐标系),
    处理 page.rotation = 90 / 180 / 270 的情形; rotation=0 直接返回。
    转换后: x = 列方向(从左到右), y = 行方向(从上到下), 与脚本其余逻辑一致。
    公式推导: rotation=90 顺时针时 mediabox (x, y) -> 视觉 (my - y, x)；
              rotation=180 旋转时 mediabox (x, y) -> 视觉 (mx - x, my - y)；
              rotation=270 顺时针时 mediabox (x, y) -> 视觉 (y, mx - x)。

    已知场景: 建行活期明细全量版(mediabox 595x842, rotation=90, 30 页横版),
    序号 1-19 视觉 y 从 142.9 递增到 526.1, 表头 9 列视觉 y=117 一行, 页脚
    y=538.3 一行, 均符合规范。
    """
    rot = page.rotation
    if rot == 0:
        return words
    mb = page.mediabox
    mx, my = mb.width, mb.height
    out = []
    for w in words:
        x0, y0, x1, y1 = w[0], w[1], w[2], w[3]
        if rot == 90:
            n0, n1 = my - y0, x0
            n2, n3 = my - y1, x1
        elif rot == 180:
            n0, n1 = mx - x0, my - y0
            n2, n3 = mx - x1, my - y1
        elif rot == 270:
            n0, n1 = y0, mx - x0
            n2, n3 = y1, mx - x1
        else:
            out.append(w)
            continue
        nx0, nx1 = min(n0, n2), max(n0, n2)
        ny0, ny1 = min(n1, n3), max(n1, n3)
        out.append((nx0, ny0, nx1, ny1, w[4], w[5], w[6], w[7]))
    return out


def log(msg, debug=False):
    if debug or not msg.startswith("[DEBUG]"):
        if PYODIDE:
            LOG_BUFFER.append(msg)
        else:
            print(msg)


# ---------- 表头与列边界检测 ----------
def detect_header_line(words):
    """在页面 words 中找出表头带。返回 (y, 表头带词列表, mode) 或 None。
    mode: 'single'(单行表头, band 仅表头行 ±2pt, 列边界不过滤) /
          'multi'(多行拆词表头, 建行活期 17 列表头词竖排拆 4-5 行,
          band 为窗口 ±24pt 并按命中列名词下界截断)。
    判定: 单行得分≥3 后, 检查表头行 ±16pt 内是否存在 x 与表头词错开(>10pt)
    的拆词段词 —— 有则 multi(拆词段 x 落在列间隙), 无则 single(数据行与表头同列)。"""
    rows = defaultdict(list)
    for w in words:
        rows[round(w[1], 1)].append(w)
    ys = sorted(rows)
    # 单行得分: 含完整表头关键词最多的行
    best_row, best_score = None, 0
    for y in ys:
        text = " ".join(w[4] for w in rows[y])
        score = sum(1 for kw in DEFAULT_HEADER_KEYWORDS if kw in text)
        if score > best_score:
            best_score, best_row = score, y
    if best_score < 3:
        return None
    row_xs = [w[0] for w in rows[best_row]]
    band16 = [w for w in words if abs(w[1] - best_row) <= 16]
    # 拆词段判据: y≠表头行 且 含中文 且 x 距表头行所有词 >10pt(落在列间隙,
    # 如建行"交易时"x=49 vs 表头"账号"19)。排除无中文词(民生时间"13:36:35"、
    # 金额等 x 在列间隙但属数据, 不是拆词段)。
    # 2026-08-19 修正: 含日期词的行是数据行, 不是表头拆词段——KA0200… 借记卡
    # 明细首条数据行(y=表头+15pt, 含 2026-08-06)被误判为拆词段 → 误入 multi
    # 模式, 整行被吸进表头带导致首条记录丢失、带内出现"余额宝"伪列。
    date_band_rows = {
        round(w[1], 1) for w in band16
        if any(p.match(w[4]) for p in DEFAULT_DATE_PATTERNS)
    }
    has_split = any(
        w[1] != best_row and round(w[1], 1) not in date_band_rows
        and re.search(r"[\u4e00-\u9fff]", w[4])
        and min(abs(w[0] - x) for x in row_xs) > 10
        for w in band16
    )
    if not has_split:
        # 单行表头: band 仅表头行 ±2pt(数据行紧贴表头 3pt, 不能扩大)
        band = [w for w in words if abs(w[1] - best_row) <= 2]
        return best_row, band, "single"
    # 多行拆词表头: ±24pt 窗口得分定位表头带中心。
    # 2026-08-19 修正: 主判据用**中心行的逐行命中度**——信息行(账号/打印时间等)
    # 与真正表头行的窗口命中数可能相同(信息行词少, 原"词数少者"并列规则会误选
    # 信息行, 如 KA0200… 借记卡明细: 表头 y=132.8 vs 信息行 y=111.8)。
    # 真正表头行含 记账日期/记账时间/余额 等完整关键词, 逐行命中度必然更高;
    # 同中心命中度时再比窗口命中度, 最后取词数少者(一页两段表取最上方)。
    # 2026-08-21 性能优化: 原实现对每个候选 y 重建 ±24pt 窗口文本并逐关键词
    # 子串扫描(O(ys²) 次 join; 建行 17 列 116 页实测 detect_header_line ~0.56s)。
    # 改为滑动窗口: 每行预计算关键词命中列表, 窗口得分 = 窗口内命中关键词的
    # 并集大小。关键词均不含空白, 不可能跨行拼接, 与"拼接后文本子串匹配"
    # 完全等价; 左右指针单调推进 → 每页 O(ys + 总命中数)。
    row_hits = [
        [kw for kw in DEFAULT_HEADER_KEYWORDS if kw in " ".join(w[4] for w in rows[y])]
        for y in ys
    ]
    best, best_ws, best_center, best_n = None, 0, 0, 10**9
    left = 0
    right = -1
    window_kw = Counter()
    window_n = 0
    for i, y in enumerate(ys):
        while left < len(ys) and ys[left] < y - 24:
            for kw in row_hits[left]:
                window_kw[kw] -= 1
                if window_kw[kw] <= 0:
                    del window_kw[kw]
            window_n -= len(rows[ys[left]])
            left += 1
        while right + 1 < len(ys) and ys[right + 1] <= y + 24:
            right += 1
            for kw in row_hits[right]:
                window_kw[kw] += 1
            window_n += len(rows[ys[right]])
        score = len(window_kw)
        center_score = len(row_hits[i])
        n_words = window_n
        better = False
        if (center_score, score) > (best_center, best_ws):
            better = True
        elif center_score == best_center and score == best_ws:
            # 同分并列: 同一表头带内(≤50pt)取词数少者(表头集中行);
            # 相距很远的两个相同表头(一页两段交易表)取最上方者,
            # 否则下方表头会把上方整段数据区切掉
            if best is not None and abs(y - best) <= 50 and n_words < best_n:
                better = True
        if better:
            best_center, best_ws, best, best_n = center_score, score, y, n_words
    if best_ws < 3:
        return None
    band = [w for w in words if best - 24 <= w[1] <= best + 24]
    # band 截断: 命中完整列名词的最大 y + 12pt(拆词段延伸), 排除紧邻的数据行
    hit_max = max((w[1] for w in band if any(n in w[4] for n in SPLIT_NAMES)), default=best)
    band = [w for w in band if w[1] <= hit_max + 12]
    return best, band, "multi"


def split_header_name(name):
    """把合并的表头词(如 '账户余额流水号')按标准列名拆分。返回列名列表"""
    hits = [n for n in SPLIT_NAMES if n in name]
    if len(hits) >= 2:
        # 去掉被更长项包含或与更长项位置重叠的子串, 保留最长项:
        # - '余额' 是 '账户余额' 的子串 → 丢弃
        # - '现转标志' 与 '标志交易金额' 在"现转标志交易金额"中重叠于"标志" → 丢弃短的
        #   (北京农商金额列保持单列; 否则跨行金额续段"00"会被拆进现转标志列漏计)
        # - '账户余额' 与 '现转标志' 在"账户余额现转标志"中仅相接不重叠 → 两列都保留
        spans = [(h, name.find(h), name.find(h) + len(h)) for h in hits]
        filtered = [
            h for h, s, e in spans
            if not any(o != h and len(o) > len(h) and s < oe and os < e
                       for o, os, oe in spans)
        ]
        return filtered
    return [name]


def detect_column_boundaries(page_words, header):
    """基于表头带词 x 聚类 + 数据词聚类, 生成列边界列表 [(lo, hi, name), ...]。
    header = (header_y, band_words, mode); 拆行表头词按 x 聚类合并(如 '交易时'+'间'→'交易时间')。
    mode='multi' 时聚类文本仅保留命中 SPLIT_NAMES 的表头词(过滤 band 内混入的数据词),
    mode='single' 时 band 只含表头行词, 不过滤(避免"交易金额/收入"等非 SPLIT_NAMES 表头词被误删)。
    返回 (cols, band_bottom_y); band_bottom_y = 表头词最大 y, 供数据区下界使用。"""
    header_y, band_words, mode = header
    if not band_words:
        return None, header_y

    if mode == "multi":
        # 多行模式: 先剔除页眉/信息行(整行含冒号, 全角"："或半角":"——民生信息区
        # "客户名称:"用半角; KA0200… 借记卡明细信息行"按收支筛选："还带无冒号词
        # "全部", 词级过滤会漏)与数据词(不含中文字符的账号/金额/日期), 再按 x 聚类。
        # 顺序关键: 若先聚类, 页眉词("-"/"20260625")会把不相干列段衔接成错误大簇。
        bad_rows = {round(w[1], 1) for w in band_words
                    if "：" in w[4] or ":" in w[4]}
        bw = sorted((w for w in band_words
                     if round(w[1], 1) not in bad_rows
                     and re.search(r"[\u4e00-\u9fff]", w[4])),
                    key=lambda w: w[0])
    else:
        bw = sorted(band_words, key=lambda w: w[0])
    clusters = []
    for w in bw:
        # 聚类判据用 x0 差(≤12pt): 表头拆词段 x0 差 4-9pt(建行"交易时"58.6→"间"67.6,
        # 中科擎云建行 17 列企业明细拆词间距 9pt; 原 8pt 阈值漏合并 → 列名"交易时"
        # 缺"间"字 → 日期词被列限定过滤 → 提取 0 条记录, 2026-08-25 修正);
        # 相邻独立列 x0 差 ≥30pt(民生"凭证类型"200→"凭证号码"235)不误聚。
        if clusters and w[0] - clusters[-1][-1][0] <= 12:
            clusters[-1].append(w)
        else:
            clusters.append([w])
    entries = []  # (x0, text, max_y)
    entry_x1 = []  # 每个表头簇的右缘 x1(与 entries 对齐), 供两级表头父表头判定
    for cl in clusters:
        cl_sorted = sorted(cl, key=lambda w: (w[1], w[0]))
        text = "".join(w[4] for w in cl_sorted)
        entries.append((cl_sorted[0][0], text, max(w[1] for w in cl_sorted)))
        entry_x1.append(max(w[2] for w in cl_sorted))
    # 2026-08-19 修正: 取消 multi 模式的 SPLIT_NAMES 列过滤。原过滤是为排除
    # band 内混入的数据词, 但 band 已被 hit_max+12 截断 + 冒号行排除双层保护
    # (数据行在表头下方 ≥12pt 且列名带冒号的信息行已被剔除), 再按 SPLIT_NAMES
    # 过滤会把合法但未登记的表头列(币别/金额/交易名称/渠道/网点名称/对方账户名,
    # 如 KA0200… 借记卡明细)误删, 导致 12 列只剩 6 列、记录按残缺列切分。
    if not entries:
        return None, header_y
    _order = sorted(range(len(entries)), key=lambda i: entries[i][0])
    entries = [entries[i] for i in _order]
    entry_x1 = [entry_x1[i] for i in _order]
    # ---- 两级表头: 剔除横跨子列的父表头词(2026-08-21, 浦发银行) ----
    # 浦发"发生额"横跨子列"借方/贷方"、"交易对手信息"横跨"对手机构/对手名称"。
    # 父表头词悬浮于两个子列之间的列间隙(与左右子列 x 均不重叠), 且其 y 在
    # 主表头行上方; 不剔除会被当成数据列, 造成 贷方金额按右对齐 x0 漂移串入
    # 账户余额列(余额被拼成 93185166.66 之类的脏值)、支出合计为 0。
    # 判据(全部满足才剔除, 保守):
    #   1) multi 模式(单行表头无上下层结构, 不适用);
    #   2) 簇 max_y 在主表头行上方 ≥4pt;
    #   3) 左右各存在一个"主行及以下"的簇, 且该簇与二者 x 均不重叠(容差 1pt)
    #      —— 父表头在子列间隙; 普通列(如交易日期, 无下级子列)左右无下方簇
    #      环绕, 不会被误删。
    if mode == "multi" and len(entries) > 2:
        keep_idx = []
        for i, (_x0, _t, my) in enumerate(entries):
            if my < header_y - 4:
                li = next((j for j in range(i - 1, -1, -1)
                           if entries[j][2] >= header_y - 4), None)
                ri = next((j for j in range(i + 1, len(entries))
                           if entries[j][2] >= header_y - 4), None)
                if (li is not None and ri is not None
                        and entry_x1[li] <= _x0 + 1.0
                        and entry_x1[i] <= entries[ri][0] - 1.0):
                    continue
            keep_idx.append(i)
        if len(keep_idx) < len(entries):
            entries = [entries[i] for i in keep_idx]
            entry_x1 = [entry_x1[i] for i in keep_idx]
    x_pos = [e[0] for e in entries]
    names = []
    for _x, t, _y in entries:
        names.extend(split_header_name(t))
    band_bottom = max(e[2] for e in entries)

    # 表头词 x → 合并词对应的多列区间
    bounds = []
    for i in range(len(x_pos)):
        lo = (x_pos[i - 1] + x_pos[i]) / 2 if i > 0 else x_pos[i] - 18
        hi = (x_pos[i] + x_pos[i + 1]) / 2 if i < len(x_pos) - 1 else x_pos[i] + 60
        bounds.append((lo, hi))

    # 数据词 x 统计(用于拆分合并列)
    # 汇总/元信息/页脚行(期末余额、服务费说明、页码行等):
    # 不是交易数据, 混入会污染日期上限与 x 聚类。
    _row_texts = {}
    for w in page_words:
        _row_texts.setdefault(round(w[1], 1), []).append(w[4])
    bad_rows = {
        y for y, ts in _row_texts.items()
        if BAL_SUMMARY_RE.search(" ".join(ts))
        or META_ROW_RE.search(" ".join(ts))
        or is_footer_word(" ".join(ts))
        or (SUMMARY_RE.search(" ".join(ts))
            and not any(any(p.match(t) for p in DEFAULT_DATE_PATTERNS) for t in ts))
    }
    # 数据区上界: 默认表头下方 400pt; 若页内能找到日期行, 收紧到"最后日期行下 8pt"。
    # 表格下方的法律声明/条款长文、侧栏说明、汇总行等都不是交易数据,
    # 若混入 x 聚类会撑乱列边界或触发误拆列。
    # 8pt 仅吸收同行的 y 取整误差; 末条记录的续行(如信用卡 REF 行)不会影响聚类
    # (续行文字与主行同列 x 范围, 主行已定义该列聚簇)。
    # 无日期行时回退 band_bottom+400(注意 max 空列表不能给 default=0, 否则
    # data_hi 会变成 8pt 导致数据窗口为空、列边界修正被跳过)
    _date_ws = [w for w in find_date_rows(page_words, DEFAULT_DATE_PATTERNS)
                if w[1] > band_bottom and round(w[1], 1) not in bad_rows]
    data_hi = band_bottom + 400.0
    if _date_ws:
        data_hi = min(data_hi, max(w[3] for w in _date_ws) + 8.0)
    # 右边缘过滤: 表格右侧的边注/说明文字
    # 会干扰 x 聚类, 排除 x0 超出表头词右缘 +40pt 的词(正常数据列不会越出表头这么多)。
    _max_hdr_x1 = max((w[2] for w in band_words), default=0.0)
    data_words = [w for w in page_words
                  if band_bottom + 5 < w[1] < data_hi and round(w[1], 1) not in bad_rows
                  and w[0] <= _max_hdr_x1 + 40
                  and "_" not in w[4]]
    cols = []
    col_anchors = []  # 每列的表头词 x 锚点(子拆分列取所在段内的表头词 x), 供边界修正守卫使用
    for i, (lo, hi) in enumerate(bounds):
        parts = split_header_name(entries[i][1])
        if len(parts) == 1:
            cols.append((lo, hi, parts[0]))
            col_anchors.append(entries[i][0])
            continue
        # 合并列: 用该区间内的数据词 x 聚类拆分
        xs = sorted({w[0] for w in data_words if lo <= w[0] < hi})
        clusters = []
        for x in xs:
            if clusters and x - clusters[-1][-1] <= 5:
                clusters[-1].append(x)
            else:
                clusters.append([x])
        if len(clusters) >= len(parts):
            # 按聚类数切分区间; 簇数 > 列名数时, 多余簇并入最左列
            # (右对齐数字列因位数不同 x0 会散成多簇, 如民生余额 452.2/459.2 同属
            # "账户余额"; 若按簇数逐段切分, 真正的右邻列(流水号 498.2)会被挤出列范围丢失)
            n_extra = len(clusters) - len(parts)
            seg_lo = lo
            for k in range(len(parts)):
                ci = min(k + n_extra, len(clusters) - 1)
                if k + 1 < len(parts):
                    nxt = min(k + 1 + n_extra, len(clusters) - 1)
                    seg_hi = (clusters[ci][-1] + clusters[nxt][0]) / 2
                else:
                    seg_hi = hi
                cols.append((seg_lo, seg_hi, parts[k]))
                col_anchors.append(entries[i][0])
                seg_lo = seg_hi
        else:
            cols.append((lo, hi, "|".join(parts)))
            col_anchors.append(entries[i][0])

    # ---- 数据词全局 x 聚类修正列边界 ----
    # 表头词可能居中而数据左对齐(如建行"摘要"表头 x=89.2 但摘要数据 x=56.1),
    # 表头中点边界会把摘要数据错分进"序号"列。用数据区前部(避开页脚)的
    # 词 x 按列间隙(>5pt 同列、列间自然分离)聚类; 聚类数 == 列数时用聚类间隙
    # 替换表头中点边界。聚类数不符(水印词干扰/列数据缺失)时回退表头中点。
    # 阈值使用表头列间距最小值的一半(min(gaps)/2), 这样在列间距差异大的
    # 对账单(如建行活期明细全量版"序号"与"摘要"列间距 24pt 但列间距最小 39pt)
    # 也能正确聚类; 列间距小时回退到 5pt(老逻辑保留)。
    data_xs = sorted(
        w[0] for w in page_words
        if band_bottom + 5 < w[1] < data_hi and w[0] <= _max_hdr_x1 + 40 and "_" not in w[4]
    )
    gaps = sorted([x_pos[i + 1] - x_pos[i] for i in range(len(x_pos) - 1)]) if len(x_pos) > 1 else []
    # 阈值 = max(5, min(表头列间距)/2 × 系数); 若聚类数 ≠ 列数, 依次放大系数重试
    # (描述列内部词间距大时, 系数放大后才能正确合并)。
    clusters_all = []
    for _factor in (1.0, 1.5, 2.0):
        t = max(5, min(gaps) / 2 * _factor) if gaps else 5
        cl = []
        for x in data_xs:
            if cl and x - cl[-1][-1] <= t:
                cl[-1].append(x)
            else:
                cl.append([x])
        if len(cl) == len(cols) or _factor == 2.0:
            clusters_all = cl
            break
    if len(clusters_all) == len(cols):
        # 守卫: 修正后的列边界必须仍位于相邻两个表头词 x 之间, 否则拒绝修正
        # (页脚/表格下方文字会混入 data_xs, 把边界拉歪)
        new_his = [(clusters_all[k][-1] + clusters_all[k + 1][0]) / 2 for k in range(len(cols) - 1)]
        valid = True
        if len(col_anchors) == len(cols):
            for k in range(len(cols) - 1):
                # 容差: 表头词居中/左对齐时, 数据聚类边界可能略超出表头词起点
                # 容忍 max(10, 20%×列间距) 的越界, 仍能拒绝页脚文字大幅拉偏边界
                tol = max(10.0, 0.3 * (col_anchors[k + 1] - col_anchors[k]))
                if not (col_anchors[k] - tol < new_his[k] < col_anchors[k + 1] + tol):
                    valid = False
                    break
        else:
            valid = False
        if valid:
            for k in range(len(cols) - 1):
                cols[k] = (cols[k][0], new_his[k], cols[k][2])
                cols[k + 1] = (new_his[k], cols[k + 1][1], cols[k + 1][2])
    return cols, band_bottom


# ---------- 微信支付交易明细专用提取 ----------
# 微信支付交易明细证明 PDF 与银行对账单布局完全不同:
#  - 竖排式记录: 每条交易占 2 行(主行: 交易单号|日期|交易类型|收/支|交易方式①|金额|交易对方①|商户单号①;
#    续行: 交易单号尾|时间|交易方式②|交易对方②|商户单号②)
#  - 仅首页有表头(y=214.9), 后续页无表头
#  - 无余额列; 金额全为正(正负由"收/支/其他"列决定)
#  - 日期格式 YYYY-MM-DD + 独立时间行 HH:MM:SS
#  - 交易单号/交易方式/交易对方/商户单号都可能跨行(主行+续行), 直接拼接
WECHAT_SIGN = "微信支付交易明细证明"
# 微信日期固定 YYYY-MM-DD; 模块级常量化(原每文件重新编译)。
# 不能用默认 4 种日期正则——商户单号的 8 位数字片段(如 20105739)会被
# ^\d{8}$ 误判为日期行, 导致续行被错误切成新记录(+1000 条)。
WECHAT_DATE_PATTERN = re.compile(r"^\d{4}-\d{1,2}-\d{1,2}$")
# (列名, x 锚点) 按 x 升序; 列边界取相邻锚点中点
WECHAT_COLS = [
    ("交易单号", 40.0), ("交易时间", 160.0), ("交易类型", 217.0),
    ("收/支/其他", 281.0), ("交易方式", 322.0), ("金额(元)", 378.0),
    ("交易对方", 422.0), ("商户单号", 470.0),
]


def wechat_column_bounds():
    xs = [c[1] for c in WECHAT_COLS]
    out = []
    for i, (name, x) in enumerate(WECHAT_COLS):
        lo = (xs[i - 1] + x) / 2 if i > 0 else 0.0
        hi = (x + xs[i + 1]) / 2 if i < len(xs) - 1 else 595.0
        out.append((lo, hi, name))
    return out


def is_wechat_doc(doc):
    """第 1 页文本含微信支付交易明细证明 → 微信专用分支"""
    try:
        return WECHAT_SIGN in doc[0].get_text("text")
    except Exception:  # noqa: BLE001
        return False


def slice_records(region, date_rows, page_footer_y, absorb=6, boundaries=None):
    """按日期行把 region 切分为记录 cluster 列表。
    用排序 + bisect 二分替代"每条记录全量扫描 region"(大文件核心性能热点):
    原逻辑等价于 cluster = [w for w in region if y_start-absorb <= w[1] < y_end-absorb],
    此处稳定排序后按 y 二分取区间, 复杂度 O(n log n) vs O(n*m)。
    absorb=记录边界吸收量(日期行上下各 absorb pt; 大间距银行如建行活期
    记录多行段跨日期 ±20pt 时由调用方传入更大的值)。
    boundaries(可选): 相邻记录的精确边界列表(长度 = len(date_rows)-1), 由
    compute_record_boundaries 对"列式多行单元格"布局(建行企业账户 17 列)计算;
    提供时: 记录 i = [boundaries[i-1] 或 region 首词, boundaries[i] 或 page_footer_y),
    不再使用固定吸收量(该布局的记录上下段可分别超出日期 20pt/24pt 甚至 +88pt,
    单值吸收无法同时覆盖上下段, 见 compute_record_boundaries 注释)。"""
    region_sorted = sorted(region, key=lambda w: w[1])
    ys = [w[1] for w in region_sorted]
    clusters = []
    for idx, dw in enumerate(date_rows):
        y_start = dw[1]
        if boundaries is not None and len(boundaries) == len(date_rows) - 1:
            lo = region_sorted[0][1] if idx == 0 else boundaries[idx - 1]
            hi = page_footer_y if idx == len(date_rows) - 1 else boundaries[idx]
            lo_i = bisect.bisect_left(ys, lo)
            hi_i = bisect.bisect_left(ys, hi)
        else:
            lo_i = bisect.bisect_left(ys, y_start - absorb)
            if idx + 1 < len(date_rows):
                hi_i = bisect.bisect_left(ys, date_rows[idx + 1][1] - absorb)
            else:
                # 最后一条记录: 下界直接取页脚, 不再减 absorb —— 否则记录底部多行单元格
                # 尾段(金额/时间/户名末行等)若越过 footer-absorb 会被整段切丢
                # (建行企业账户 17 列明细页尾记录底部可达页脚上方 12pt 内)。
                hi_i = bisect.bisect_left(ys, page_footer_y)
        if lo_i < hi_i:
            clusters.append(region_sorted[lo_i:hi_i])
    return clusters


def _nearest_date_idx(date_ys, y0):
    """按 bisect 找最近日期行下标(并列取较早者, 与 min(range, key=abs) 语义一致)。
    2026-08-21 性能优化: 原 compute_record_boundaries / compute_longtext_cells 对每个
    运行段都做 min(range, key=abs) 全量扫描(建行 17 列 116 页约 11 万次调用, ~0.4s),
    bisect 二分 O(log n) 且语义完全等价。"""
    i = bisect.bisect_left(date_ys, y0)
    if i == 0:
        return 0
    if i >= len(date_ys):
        return len(date_ys) - 1
    return i - 1 if y0 - date_ys[i - 1] <= date_ys[i] - y0 else i


def compute_record_boundaries(region, date_rows, cols, debug=False):
    """对"列式多行单元格"布局计算相邻记录的精确 y 边界。

    背景(建行企业账户 17 列明细, 2026-08-18): 每条记录的各列单元格被 PDF 拆成
    2-6 行纵向堆叠, 记录上下段可分别超出日期锚点 20pt/24pt(个别税务记录备注
    续行达 +88pt, 对应记录间距 108pt)。固定吸收量无法同时覆盖上段与下段:
    - 吸收 ≥20pt 才能含顶部首行(如 华夏银 在日期上方 20pt);
    - 吸收 < 间距-下延 才能不切到下段(如 40pt 间距下段 +20pt 时吸收须 <20pt)。
    两者互相矛盾 → 必须按页实际词布局逐对计算边界。

    方法: 每个列内, 同一记录的单元格词按 y 以 8pt 行距连续排列, 跨记录间隙
    ≥12pt → 按间隙 >10pt 把列内词切成"运行段"; 每个运行段按起点 y 归属最近
    的日期行(记录上段起点总在自身日期附近, 长备注续行起点仍离自身日期更近);
    边界 i = (记录 i 最大词 y + 记录 i+1 最小词 y) / 2。

    长文本列(备注/企业流水号等, 单元格可纵向与下一条记录首行重叠, 如备注
    "收费项目:对公人民币转账、汇款…" 续行、企业流水号 "…%#^支付机构交易流水号
    ：20250314…" 纵跨两记录行带)不参与边界计算——这些列由 compute_longtext_cells
    按列内运行段独立归属。本函数迭代识别所有"造成记录 y 重叠"的列并排除。

    返回: (边界列表, 排除列下标集合); 若归属不稳定(记录无词/边界越界/非单调)
    返回 (None, 排除列集合), 调用方回退固定吸收逻辑。"""
    if len(date_rows) < 2:
        return None, set()
    excluded = {i for i, (_lo, _hi, nm) in enumerate(cols) if "备注" in nm}
    for _attempt in range(8):
        col_words = defaultdict(list)
        for w in region:
            ci = match_column(w[0], w[4], cols)
            if ci is not None and ci not in excluded:
                col_words[ci].append(w)
        runs = []  # (运行段起点 y, [words])
        for ci, ws in col_words.items():
            ws_sorted = sorted(ws, key=lambda w: (w[1], w[0]))
            cur = []
            prev_y = None
            for w in ws_sorted:
                if cur and w[1] - prev_y > 10:
                    runs.append((cur[0][1], list(cur)))
                    cur = []
                cur.append(w)
                prev_y = w[1]
            if cur:
                runs.append((cur[0][1], list(cur)))
        if not runs:
            return None, excluded
        # 运行段归属: 按起点 y 找最近的日期行
        record_words = defaultdict(list)
        date_ys = [d[1] for d in date_rows]
        for y0, ws in runs:
            di = _nearest_date_idx(date_ys, y0)
            record_words[di].extend(ws)
        if any(not record_words[i] for i in range(len(date_rows))):
            log("[DEBUG] 记录边界计算: 存在无词记录, 回退吸收法", debug)
            return None, excluded
        ys_min = [min(w[1] for w in record_words[i]) for i in range(len(date_rows))]
        ys_max = [max(w[1] for w in record_words[i]) for i in range(len(date_rows))]
        bad_pair = None
        for i in range(len(date_rows) - 1):
            if ys_max[i] >= ys_min[i + 1]:
                bad_pair = i
                break
        if bad_pair is None:
            boundaries = []
            for i in range(len(date_rows) - 1):
                b = (ys_max[i] + ys_min[i + 1]) / 2.0
                if not (date_rows[i][1] <= b <= date_rows[i + 1][1]):
                    log(f"[DEBUG] 记录边界计算: 边界 {b:.1f} 越出日期区间, 回退吸收法", debug)
                    return None, excluded
                boundaries.append(b)
            if any(boundaries[i] >= boundaries[i + 1] for i in range(len(boundaries) - 1)):
                log("[DEBUG] 记录边界计算: 边界非单调, 回退吸收法", debug)
                return None, excluded
            return boundaries, excluded
        # 重叠对: 只有"上一条记录的单元格向下延伸越过下一条记录首行"才造成
        # 重叠(长文本列, 如企业流水号/对方户名), 下一条记录的正常首行词不是
        # 原因 → 仅排除上一条记录中越界词所在的列, 避免把整页列全部排除。
        overlap_cols = set()
        for w in record_words[bad_pair]:
            if w[1] >= ys_min[bad_pair + 1]:
                ci = match_column(w[0], w[4], cols)
                if ci is not None:
                    overlap_cols.add(ci)
        if not overlap_cols or overlap_cols <= excluded:
            log(f"[DEBUG] 记录边界计算: 记录 {bad_pair}/{bad_pair+1} y 重叠且无法排除列, 回退吸收法", debug)
            return None, excluded
        excluded |= overlap_cols
    log("[DEBUG] 记录边界计算: 排除列迭代超限, 回退吸收法", debug)
    return None, excluded


def compute_longtext_cells(region, date_rows, cols, excluded, debug=False):
    """按列内运行段独立重建长文本列(备注/企业流水号等)的单元格。

    长文本单元格纵向可与下一条记录首行重叠(备注续行、企业流水号长串等),
    纯 y 切片会截断/串行。方法: 列内按间隙 >10pt 切运行段, 每段归属
    "最后一个 ≤ (段起点+30pt) 的日期行"。理由: 长文本段可落在自身记录日期
    上方 8-16pt(如"十里堡/财务一季度服务费"始于 d-8/d-4)或下方 28-52pt
    (被上一条长单元格推后), 起点本身相对日期漂移 ±52pt, 单用"≤起点"会把
    d-8 段误当页首残词、把 d+52 段误归下一条记录; +30pt 后两种形态都落在
    自身记录日期与下一条记录日期之间(实测间距 ≥36pt)。空单元格记录自然跳过;
    起点早于首日期 30pt 以上的词段视为页首固定词(表头残词)跳过。

    返回 {列下标: {记录索引: [按 y 排序的 (y, text)]}}; 若归属顺序退化
    (段归属记录号回退)返回 {} 让调用方沿用 y 切片结果。"""
    if len(date_rows) < 2:
        return {}
    out = {}
    for ci in sorted(excluded):
        ws = sorted((w for w in region if match_column(w[0], w[4], cols) == ci),
                    key=lambda w: (w[1], w[0]))
        if not ws:
            continue
        runs = []
        cur = []
        prev_y = None
        for w in ws:
            if cur and w[1] - prev_y > 10:
                runs.append(list(cur))
                cur = []
            cur.append(w)
            prev_y = w[1]
        if cur:
            runs.append(list(cur))
        cell = {}
        last_ri = -1
        ok = True
        date_ys = [d[1] for d in date_rows]
        for run in runs:
            y0 = run[0][1]
            # 归属规则(2026-08-21 修正, 保守): 原规则"最后一个 ≤ y0+30pt 的日期行"
            # 在**窄行距**下会越过下一条日期——浦发电子对账单(行距 26pt)2 行备注
            # 单元格第二行在日期下方 5-6pt, y0+30 越过下一条日期行 → 备注整列
            # 串行; 建行 17 列(行距 ≥36pt)长备注续行被推后 +28~+52pt 必须依赖
            # +30 规则(最近日期反而是下一条, 会串行)。
            # 判定: 仅当"行距 <30pt 且 段起点距最近日期行 ≤8pt"(浦发特征)时改用
            # 最近日期行归属, 其余场景完全保留 +30 原规则, 不影响建行 17 列等。
            nearest = _nearest_date_idx(date_ys, y0)
            d_near = abs(date_ys[nearest] - y0)
            pitch = (date_ys[nearest + 1] - date_ys[nearest]
                     if nearest + 1 < len(date_rows) else 10**9)
            if d_near <= 8.0 and pitch < 30.0:
                ri = nearest
            else:
                ri = bisect.bisect_right(date_ys, y0 + 30) - 1
            if ri < 0:
                continue  # 页首固定词段(如表头残词), 非数据
            if ri < last_ri:
                log(f"[DEBUG] 长文本列归属退化(列{ci} ri={ri} < {last_ri}), 放弃逐列重建", debug)
                ok = False
                break
            last_ri = ri
            cell.setdefault(ri, []).extend((w[1], w[4]) for w in sorted(run, key=lambda x: (x[1], x[0])))
        if ok and cell:
            out[ci] = cell
    return out


def is_multiline_column_layout(region, date_rows, cols):
    """判断页面是否为"列式多行单元格"布局(建行企业账户 17 列明细等):
    多数记录的日期锚点上方(16pt 内)存在 ≥2 个非日期列的词且分布在 ≥2 个不同
    y 层(记录首行 账号/对方户名/对方账号/对方开户机构/流水号 等段落在日期行
    上方 4-16pt, 且因列式错峰分布在不同行高)。普通单行/双行对账单不满足:
    - 民生/邮储/招商/兴业/顺义/北京农商/建行 hqmx 等日期上方至多 1 个词;
    - 旅立方(2 行式)日期上方虽有上一条记录的续行词, 但续行在日期上方 21-25pt
      (超出 16pt 窗口), 且属于"上一条记录"而非本记录的首行 → 不误判。
    窗口取 16pt: 17 列布局首行最远 -20pt(华夏银), 但同记录其它首行(-8~-12pt)
    仍在窗口内, 满足 ≥2 词/≥2 层; 旅立方上一条续行 -21pt 起全部排除。"""
    if len(date_rows) < 2:
        return False
    multi = 0
    for dw in date_rows:
        above = []
        for w in region:
            if dw[1] - 16 <= w[1] < dw[1] - 0.5:
                ci = match_column(w[0], w[4], cols)
                if ci is not None and not any(k in cols[ci][2] for k in ("日期", "时间")):
                    above.append(w)
        if len(above) >= 2 and len({round(w[1], 1) for w in above}) >= 2:
            multi += 1
    return multi >= max(1, int(len(date_rows) * 0.5))


def _centered_mis_slice(region, date_rows, cols, absorb=6.0):
    """居中多行单元格布局的"切片错位词"计数(2026-08-21, 浦发银行电子对账单)。

    浦发(页面 rotation=90 横版)的单元格以记录行带**垂直居中**: 3-5 行单元格的
    首行可高出日期行 11-17pt(如 对手名称"京蝎府（北京）酒店"在日期上方 12pt),
    固定 absorb=6 的 slice_records 切片会把首行切给**上一条**记录(对手名称缺
    前缀、上一条记录被污染)。而普通银行(民生/邮储/兴业/金小九等)日期上方的词
    是**上一条记录的尾部**(距上一条日期更近), absorb=6 切片天然正确 → 计数 0。

    判据: 统计"落入切片 j(吸收法)但距日期行 j+1 比 j 更近"的非日期列词数
    (即 w.y > 相邻两日期行中点, 切片归属与最近日期行归属不一致)。实测区分度:
    浦发 2-4 词/页(所有页一致), 民生/旅立方/北京银行/hqmx/兴业/金小九 0 词/页,
    华夏(交易机构绝对定位段)11-31 词/页、建行 17 列 30+ 词/页(后者本就由
    _sig17 门控走精确边界, 不受影响)。"""
    if len(date_rows) < 2:
        return 0
    dys = [d[1] for d in date_rows]
    n = 0
    for w in region:
        ci = match_column(w[0], w[4], cols)
        if ci is None:
            continue
        nm = cols[ci][2]
        if "日期" in nm or "时间" in nm:
            continue
        # 词所在切片 j: [dys[j]-absorb, dys[j+1]-absorb)
        j = bisect.bisect_right(dys, w[1]) - 1
        if j < 0 or j >= len(dys) - 1:
            continue
        if not (dys[j] - absorb <= w[1] < dys[j + 1] - absorb):
            continue
        if w[1] > (dys[j] + dys[j + 1]) / 2:
            n += 1
    return n


def _cluster_edges(vals, tol=1.0):
    """把浮点坐标边按 tol 容差聚簇(如 77.6/77.7 表头与数据格微差合并)。"""
    out = []
    for v in sorted(vals):
        if not out or v - out[-1] > tol:
            out.append(v)
    return out


def detect_table_grid(rects):
    """从 PDF 绘图矩形识别"完整单元格网格"。返回 (列边列表, 行边列表) 或 None。

    部分银行对账单(建行企业账户 17 列等)用**填充矩形**逐格绘制表格(每格一个
    白底黑框矩形): 列边 = 所有矩形 x0/x1 聚类, 行边 = 所有矩形 y0/y1 聚类,
    网格直接给出精确列边界与"一行=一条记录"的行带。

    门控(避免误伤只有表框/分隔线的格式):
    - 只统计"填充单元格矩形"(fill, 宽高均 >5pt)——民生/邮储/微信/华夏/旅立方/
      hqmx49 只有描边线(或招商高 0.1pt 的细填充条), 计数为 0 直接返回 None;
    - 列边 ≥3 条(≥2 列)且绝大多数矩形四边对齐网格;
    - 矩形数量接近 列数×行数(允许少量合并单元格)。
    """
    seen = set()
    cells = []
    for d in rects:
        r = d["rect"]
        key = (round(r.x0, 2), round(r.y0, 2), round(r.x1, 2), round(r.y1, 2))
        if key in seen:
            continue  # 同一格常同时含 fill 与 stroke, 去重
        seen.add(key)
        if d.get("type") in ("f", "s") and r.width > 5 and r.height > 5:
            cells.append(r)
    if len(cells) < 8:
        return None
    vv = _cluster_edges({v for r in cells for v in (r.x0, r.x1)})
    hh = _cluster_edges({v for r in cells for v in (r.y0, r.y1)})
    if len(vv) < 4 or len(hh) < 2:
        return None
    aligned = sum(
        1 for r in cells
        if min(abs(r.x0 - e) for e in vv) <= 1.0
        and min(abs(r.x1 - e) for e in vv) <= 1.0
        and min(abs(r.y0 - e) for e in hh) <= 1.0
        and min(abs(r.y1 - e) for e in hh) <= 1.0
    )
    if aligned / len(cells) < 0.9:
        return None
    expected = (len(vv) - 1) * (len(hh) - 1)
    if len(cells) < expected * 0.6:
        return None  # 覆盖太少 → 不是完整单元格网格
    return vv, hh


def grid_row_bands(hh, header_y, page_footer_y, words=None, header_keywords=None):
    """网格行带: 表头行之下的行区间 [(y0,y1), ...], 每行 = 一条记录。

    表头带判别(2026-08-29 修正): 优先按"行带内文本是否含表头关键词"判定 ——
    仅首页有表头的格式(浦发企业回单/浦发个人流水), 后续页网格里根本没有表头带,
    若沿用首页 header_y 过滤, 该 y 会落进后续页某条数据行带, 把它误当表头带
    剔除 → 行带数 != 记录数 → 网格路径整体失效。words 提供时按文本判别,
    未提供时回退旧的 y 匹配(兼容既有调用)。"""
    bands = []
    for i in range(len(hh) - 1):
        y0, y1 = hh[i], hh[i + 1]
        if y0 >= page_footer_y - 0.5:
            continue  # 页脚之下的残留带
        if words is not None and header_keywords:
            # 文本判别: 带内含 >=2 个表头关键词才算表头带
            hits = sum(
                1 for k in header_keywords
                if any(k in w[4] for w in words if y0 - 0.5 <= (w[1] + w[3]) / 2 < y1 + 0.5)
            )
            if hits >= 2:
                continue  # 表头带
        elif y0 <= header_y < y1:
            continue  # 表头带(旧 y 匹配回退)
        bands.append((y0, y1))
    return bands

def split_cross_column_words(page, words, cols, margin=1.0):
    """网格路径专用: 把横向跨越列边界的粘连词按字符级坐标切分回各列。

    场景(浦发个人流水): "王毛敖海621700001"在 words 层是一个词(x=379-461,
    横跨对手姓名列 377-418.8 与对手账号列 418.8-464.9), 词级 match_column
    按起点 x 整词归入姓名列 → 账号丢前 9 位。rawdict 给出每字符精确 bbox:
    '王毛敖海' x=379-415(姓名列), '621700001' x=420.8-461.3(账号列),
    按字符中心点落列即可精确切分。

    规则: 词起点列 != 词终点列(终点按词 x1-ε 定位)才切分; 单列词原样保留。
    返回新词列表(tuple 结构与 pymupdf words 兼容: 前 5 元素为
    (x0,y0,x1,y1,text), 其余字段从原词继承)。"""
    def _col_at(x):
        for i, (lo, hi, _) in enumerate(cols):
            if lo - margin <= x < hi + margin:
                return i
        return None

    out = []
    need_raw = False
    for w in words:
        c0 = _col_at(w[0])
        c1 = _col_at(w[2] - 0.01)
        if c0 is None or c1 is None or c0 == c1:
            out.append(w)
            continue
        need_raw = True
        out.append(None)  # 占位, 待字符级切分
    if not need_raw:
        return out
    try:
        raw = page.get_text("rawdict")
    except Exception:  # noqa: BLE001
        return [w for w in out if w is not None] or list(words)
    all_chars = []
    for block in raw.get("blocks", []):
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                for ch in span.get("chars", []):
                    b = ch["bbox"]
                    all_chars.append((b[0], b[1], b[2], b[3], ch["c"]))
    result = []
    for idx, w in enumerate(out):
        if w is not None:
            result.append(w)
            continue
        w = words[idx]  # 占位 → 原词
        x0, y0, x1, y1, text = w[0], w[1], w[2], w[3], w[4]
        sub = [c for c in all_chars
               if y0 - 1 <= (c[1] + c[3]) / 2 <= y1 + 1 and x0 - 0.5 <= c[0] <= x1 + 0.5]
        picked = []
        pool = sorted(sub, key=lambda c: (round(c[1], 1), c[0]))
        ti = 0
        for c in pool:
            if ti < len(text) and c[4] == text[ti]:
                picked.append(c)
                ti += 1
        if not picked:
            result.append(w)  # 匹配失败原样保留(归起点列)
            continue
        by_col = {}
        for c in picked:
            cx = (c[0] + c[2]) / 2
            ci = _col_at(cx)
            if ci is None:
                ci = _col_at(c[0])
            if ci is None:
                ci = _col_at(w[0])
            by_col.setdefault(ci, []).append(c)
        for ci, cs in sorted(by_col.items()):
            cs.sort(key=lambda c: c[0])
            txt = "".join(c[4] for c in cs)
            if not txt:
                continue
            nx0 = min(c[0] for c in cs)
            nx1 = max(c[2] for c in cs)
            # 继承原词第 5 位之后的字段(块/行/词号), 保持词元组结构
            tail_fields = tuple(w[5:]) if len(w) > 5 else ()
            result.append((nx0, y0, nx1, y1, txt) + tail_fields)
    return result
def _norm_anchor(s):
    """锚点文本归一化: 去空白/全角空格(视觉模型输出可能带空格)。"""
    return s.replace(" ", "").replace("\u3000", "").replace("\n", "").replace("\t", "")


def _norm_anchor(s):
    """锚点文本归一化: 去空白/全角空格(视觉模型输出可能带空格)。"""
    return s.replace(" ", "").replace("\u3000", "").replace("\n", "").replace("\t", "")


def _anchor_row_scores(row_text, normed):
    """单行对锚点的命中度: (完整锚点子串命中数, 锚点字符覆盖率之和)。
    字符覆盖率用于表头词被拆得极碎(建行 17 列"交易时/间"、"借方发/生额")的场景。
    只统计 CJK 字符: 锚点"对方卡/账号"的 "/" 不能因英文行 "Card/Account"
    的斜杠而命中(华夏英文表头行), 覆盖率信号只应由中文字符提供。"""
    row_chars = set(re.findall(r"[\u4e00-\u9fff]", row_text))
    full = 0
    char = 0.0
    for a in normed:
        if a and a in row_text:
            full += 1
        cs = set(re.findall(r"[\u4e00-\u9fff]", a))
        if cs:
            char += len(cs & row_chars) / len(cs)
    return full, char


def find_header_band_by_anchors(words, anchors, debug=False):
    """表头检测失败时(特殊列名/双语/盖章遮挡), 用视觉锚点文本反推表头带。

    返回 (header_y, band_words) 兼容 ground_header_anchors; 找不到返回 None。
    思路:
      1. 逐行算"锚点命中度"(完整子串 + 字符覆盖率), 取命中度最高的行为中心
         (信息行/数据行的命中度远低于表头行, 不会误选);
      2. 从中心行向上下扩展: 相邻行 y 间隙 ≤10pt 且与锚点相关(完整命中 /
         字符覆盖率 ≥0.7 / 与已入带词 x 重叠且覆盖率 ≥0.15)才并入带,
         遇到数据行(y 间隙 >10pt 或无关)即停 → 表头带不吸入首条数据行。
    例: 民生 表头 y=97.8, 上方信息行 y=80.3(间隙 17.5 停)、下方数据 y=111.3(停);
         北京农商 3 行表头 y=96/101/106, 数据首行 y=117.4(间隙 11.4>10 停);
         建行25 表头 y=65/73/77/81/89, 数据首行 y=101(间隙 12>10 停);
         华夏 中文表头 y=202.3 + 英文表头 y=211.6(英文行无锚点相关, 不入带,
         但英文行位于首条记录吸收窗口上方, 提取时自然丢弃, 不影响结果)。"""
    normed = [a for a in (_norm_anchor(a) for a in anchors) if a]
    if not normed:
        return None
    rows = defaultdict(list)
    for w in words:
        rows[round(w[1], 1)].append(w)
    ys = sorted(rows)
    if not ys:
        return None  # 无文字层(扫描件/纯图片页)
    scores = {}
    for y in ys:
        f, c = _anchor_row_scores("".join(w[4] for w in rows[y]), normed)
        scores[y] = (f, c)
    best_y = max(ys, key=lambda y: scores[y][0] + 0.5 * scores[y][1])
    if scores[best_y][0] + 0.5 * scores[best_y][1] < 1.5:
        return None
    # 从中心行扩展表头带(按 y 间隙 + 锚点相关性)
    in_band = {round(best_y, 1)}
    span_lo = span_hi = round(best_y, 1)
    changed = True
    while changed:
        changed = False
        for y in ys:
            if y in in_band:
                continue
            if abs(y - span_lo) <= 10 or abs(y - span_hi) <= 10:
                f, c = scores[y]
                x_overlap = any(
                    abs(w[0] - v[0]) <= 8
                    for w in rows[y] for ey in in_band for v in rows[ey])
                if f >= 1 or c >= 0.7 or (x_overlap and c >= 0.15):
                    in_band.add(y)
                    span_lo = min(span_lo, y)
                    span_hi = max(span_hi, y)
                    changed = True
    band = [w for w in words if round(w[1], 1) in in_band]
    if not band:
        return None
    # 返回带内最小 y 作为 header_y(与 detect_header_line 多行模式一致):
    # ground_header_anchors 内部按 header_y-4 裁剪带, 若返回中心行会把
    # 上方拆词行(建行25 "账户明/交易时/借方发" y=65/73, 中心行 77)切掉。
    header_y = min(w[1] for w in band)
    if debug:
        log(f"[DEBUG] 锚点反推表头带: y={header_y:.1f} 行={sorted(in_band)} band={len(band)}词", debug)
    return header_y, band


def ground_header_anchors(words, anchors, header_y, header_band, debug=False):
    """视觉锚点反查(2026-08-19): 给定锚点文本(视觉模型读出的表头列名),
    在文字层表头带内反查定位。返回按 x 升序的 [(锚点原文, 规范列名, x0, x1), ...],
    命中锚点过少或顺序异常时返回 None(调用方应回退启发式列识别)。

    **列名取自文字层而非锚点文本**: 视觉读出的锚点可能有错字/漏字(如"凭证类"
    缺"种")。锚点只负责定位列, 列名用命中簇在文字层中的拼接文本(如
    "账户明細编号-交易流水号"→"账户明细编号-交易流水号"), 保证下游依赖规范
    列名的逻辑(17 列签名门控、日期列识别、统计)不受视觉误读影响。

    拆词处理: 表头词常被 PDF 拆成多段且纵向错位(交易时/间、借方发/生额、
    账户明/细编号-/交易流/水号、凭证种/类), 无法按行内拼接匹配 → 先把表头带
    词按 x0 聚类(容差 8/12/16pt 依次尝试)还原"候选列", 在列内按 y 拼接后
    匹配锚点, 用命中词的外接盒作为该列 x 区间。

    **合并表头处理(2026-08-19 增强)**:
    - 同一单词内含多个锚点且区间不重叠(民生"账户余额流水号"→ 账户余额/流水号)
      → 按字符偏移把单词 x 区间比例拆成多列;
    - 多锚点落在不同词(北京农商"现转/标志交易金额"为同一列的拆行词)
      → 合并为一列(簇的外接盒 + 簇拼接文本)。

    header_y/header_band: detect_header_line 返回的表头行 y 与表头带词列表。
    表头带是 ±24pt 窗口, 会把页眉"本方户名/打印日期"(y=50-52)也包进来 → 按
    y ≥ header_y-4 裁剪, 只保留真正的表头行(65-89), 否则这些页眉词会污染
    候选列聚类(如"打印日期"与"凭证种类"x 相近被并入同一簇)。
    """
    band = [w for w in words
            if header_y - 4 <= w[1] <= max(w0[1] for w0 in header_band) + 2]
    if not band:
        return None
    normed = {a: _norm_anchor(a) for a in anchors if _norm_anchor(a)}
    for tol in (8.0, 12.0, 16.0):
        clusters = []
        for w in sorted(band, key=lambda x: (x[0], x[1])):
            if clusters and w[0] - clusters[-1][0][0] <= tol:
                clusters[-1].append(w)
            else:
                clusters.append([w])
        matched = []
        for a, na in normed.items():
            hit = None
            for ci, c in enumerate(clusters):
                ordered = sorted(c, key=lambda w: w[1])
                txt = "".join(w[4] for w in ordered)
                offs, pos = [], 0
                for w in ordered:
                    offs.append(pos)
                    pos += len(w[4])
                # 1) 单个词精确等于锚点
                for k, w in enumerate(ordered):
                    if w[4] == na:
                        hit = (ci, offs[k], offs[k] + len(na), [k])
                        break
                if hit is not None:
                    break
                # 2) 候选列内按 y 拼接后包含锚点
                i = txt.find(na)
                if i >= 0:
                    j = i + len(na)
                    sel = [k for k, (o, w) in enumerate(zip(offs, ordered))
                           if o < j and o + len(w[4]) > i]
                    hit = (ci, i, j, sel)
                    break
            if hit is None:
                # 3) 单字投票: 锚点字符在哪个候选列命中最多, 该列即锚点所在列
                chars = list(dict.fromkeys(na))
                best_ci, best_score = None, 0
                for ci, c in enumerate(clusters):
                    score = sum(1 for ch in chars if ch in "".join(w[4] for w in c))
                    if score > best_score:
                        best_ci, best_score = ci, score
                need = max(2, int(len(chars) * 0.6 + 0.999))
                if best_ci is not None and best_score >= need:
                    ordered = sorted(clusters[best_ci], key=lambda w: w[1])
                    hit = (best_ci, 0, len("".join(w[4] for w in ordered)),
                           list(range(len(ordered))))
            if hit is not None:
                matched.append((a, hit))
        if not matched:
            continue
        # 按簇分组, 处理"合并表头词拆列/多锚点同列合并"
        by_cluster = {}
        for a, (ci, i, j, sel) in matched:
            by_cluster.setdefault(ci, []).append((a, i, j, sel))
        cols_raw = []
        band_bottom = 0.0
        for ci, items in by_cluster.items():
            ordered = sorted(clusters[ci], key=lambda w: w[1])
            band_bottom = max(band_bottom, max(w[1] for w in ordered))
            offs, pos = [], 0
            for w in ordered:
                offs.append(pos)
                pos += len(w[4])
            txt = "".join(w[4] for w in ordered)
            items.sort(key=lambda t: t[1])
            if len(items) == 1:
                a, i, j, sel = items[0]
                ws = [ordered[k] for k in sel]
                # 列名取文字层"完整命中词"跨度: 视觉锚点可能少读末字(如
                # "账户明细编号-交易流水" 缺"号"), 用 sel 词的完整 span 还原
                # 文字层全名("账户明细编号-交易流水号"), 避免 17 列签名/统计
                # 因列名截断而失效。
                if sel:
                    lo = min(offs[k] for k in sel)
                    hi = max(offs[k] + len(ordered[k][4]) for k in sel)
                else:
                    lo, hi = i, j
                name = "|".join(split_header_name(txt[lo:hi]))
                cols_raw.append((a, name, min(w[0] for w in ws), max(w[2] for w in ws)))
                continue
            # 多锚点同簇: 全部落在同一单词内且区间不重叠 → 合并表头词拆列
            same_word = all(len(t[3]) == 1 and t[3][0] == items[0][3][0] for t in items)
            non_overlap = all(items[k][2] <= items[k + 1][1] for k in range(len(items) - 1))
            if same_word and non_overlap:
                k = items[0][3][0]
                w = ordered[k]
                w_start = offs[k]
                w_len = len(w[4])
                for a, i, j, _sel in items:
                    x0 = w[0] + (w[2] - w[0]) * (i - w_start) / w_len
                    x1 = w[0] + (w[2] - w[0]) * (j - w_start) / w_len
                    # 相邻拆分列边界各让 0.5pt, 避免恰好相切触发重叠校验
                    # (中点边界仍落在原切分点, 不影响列划分)
                    if i > w_start:
                        x0 += 0.5
                    if j < w_start + w_len:
                        x1 -= 0.5
                    cols_raw.append((a, "|".join(split_header_name(txt[i:j])), x0, x1))
                continue
            # 多行同列(如北京农商"现转/标志交易金额")→ 合并为一列
            a0 = items[0][0]
            name = "|".join(split_header_name(txt))
            cols_raw.append((a0, name, min(w[0] for w in ordered), max(w[2] for w in ordered)))
        if len(cols_raw) < max(3, len(normed) // 2 + 1):
            continue
        cols_raw.sort(key=lambda m: m[2])
        if any(cols_raw[i][3] >= cols_raw[i + 1][2] for i in range(len(cols_raw) - 1)):
            continue
        if debug:
            log(f"[DEBUG] 锚点反查: 命中 {len(cols_raw)} 列/{len(normed)} 锚点 (tol={tol:g})", debug)
        return cols_raw, band_bottom
    return None


def ground_anchors_on_page(words, anchors, debug=False):
    """视觉兜底: 无需 detect_header_line, 直接用锚点反推表头带并反查列。

    返回 (header_y, band_words, matched, band_bottom) 或 None。
    用于: P0-1 表头识别失败自动视觉兜底、P2-5 微信动态列锚点、
    P0-2 格式描述符的锚点反查(表头特殊时)。"""
    found = _anchor_page_header(words, anchors, debug=debug)
    if found is None:
        return None
    header_y, band, band_bottom = found
    cols_raw, _bb = ground_header_anchors(words, anchors, header_y, band, debug=debug)
    return header_y, band, cols_raw, band_bottom


def _anchor_page_header(words, anchors, debug=False):
    """锚点反推单页表头带: 返回 (header_y, band_words, band_bottom) 或 None。
    ground_anchors_on_page / extract_statement(每页兜底) 共用。"""
    found = find_header_band_by_anchors(words, anchors, debug=debug)
    if found is None:
        return None
    header_y, band = found
    matched = ground_header_anchors(words, anchors, header_y, band, debug=debug)
    if matched is None:
        return None
    _cols_raw, band_bottom = matched
    return header_y, band, band_bottom


def build_columns_from_anchors(matched):
    """由锚点盒(按 x 升序)构造列模板: 内部边界 = 相邻盒中点,
    首列下界 -18pt、末列上界 +60pt(与启发式列边界约定一致)。"""
    matched = sorted(matched, key=lambda m: m[2])
    x0s = [m[2] for m in matched]
    cols = []
    for i, (_a, name, x0, x1) in enumerate(matched):
        # 与启发式 detect_column_boundaries 完全一致: 边界 = 相邻列表头词
        # x0(左缘)中点。不能用 x1——华夏等"表头窄/数据宽"格式下 x1 中点会把
        # 左对齐的下一列数据(如摘要 E商宝 x=77.3)吞进上一列。
        lo = x0 - 18.0 if i == 0 else (x0s[i - 1] + x0) / 2.0
        hi = x0 + 60.0 if i == len(matched) - 1 else (x0 + x0s[i + 1]) / 2.0
        cols.append((lo, hi, name))
    return cols


def refine_cols_with_data_x(page_words, cols, band_bottom):
    """数据词全局 x 聚类修正列边界(2026-08-19, 视觉锚点模板专用)。

    锚点中点边界只基于表头盒, 对"表头居中/数据左对齐"或"表头窄/数据宽"的格式
    (邮储/华夏/hqmx49 等)会把相邻列数据错分。用数据区词 x 按列间隙聚类,
    聚类数 == 列数时用聚类间隙替换边界; 聚类数不符或修正边界越出相邻列锚点时
    拒绝修正。逻辑与 detect_column_boundaries 的数据聚类修正一致。
    """
    if len(cols) < 2:
        return cols
    _row_texts = {}
    for w in page_words:
        _row_texts.setdefault(round(w[1], 1), []).append(w[4])
    bad_rows = {
        y for y, ts in _row_texts.items()
        if BAL_SUMMARY_RE.search(" ".join(ts))
        or META_ROW_RE.search(" ".join(ts))
        or is_footer_word(" ".join(ts))
        or (SUMMARY_RE.search(" ".join(ts))
            and not any(any(p.match(t) for p in DEFAULT_DATE_PATTERNS) for t in ts))
    }
    _date_ws = [w for w in find_date_rows(page_words, DEFAULT_DATE_PATTERNS)
                if w[1] > band_bottom and round(w[1], 1) not in bad_rows]
    data_hi = band_bottom + 400.0
    if _date_ws:
        data_hi = min(data_hi, max(w[3] for w in _date_ws) + 8.0)
    _max_hdr_x1 = max(c[1] for c in cols)
    data_xs = sorted(w[0] for w in page_words
                     if band_bottom + 5 < w[1] < data_hi and round(w[1], 1) not in bad_rows
                     and w[0] <= _max_hdr_x1 + 40 and "_" not in w[4])
    anchors = [(c[0] + c[1]) / 2 for c in cols]
    gaps = sorted(anchors[i + 1] - anchors[i] for i in range(len(anchors) - 1))
    clusters_all = []
    for factor in (1.0, 1.5, 2.0):
        t = max(5, min(gaps) / 2 * factor) if gaps else 5
        cl = []
        for x in data_xs:
            if cl and x - cl[-1][-1] <= t:
                cl[-1].append(x)
            else:
                cl.append([x])
        if len(cl) == len(cols) or factor == 2.0:
            clusters_all = cl
            break
    if len(clusters_all) == len(cols):
        new_his = [(clusters_all[k][-1] + clusters_all[k + 1][0]) / 2
                   for k in range(len(cols) - 1)]
        valid = True
        for k in range(len(cols) - 1):
            tol = max(10.0, 0.3 * (anchors[k + 1] - anchors[k]))
            if not (anchors[k] - tol < new_his[k] < anchors[k + 1] + tol):
                valid = False
                break
        if valid:
            for k in range(len(cols) - 1):
                cols[k] = (cols[k][0], new_his[k], cols[k][2])
                cols[k + 1] = (new_his[k], cols[k + 1][1], cols[k + 1][2])
    return cols


def _wechat_data_anchors(words, header_y=214.9, max_rows=12):
    """微信数据列 x0 锚点: 以硬编码 WECHAT_COLS 锚点为先验, 在前 N 条主行
    (含日期词的行)中找距每个锚点最近的词 x0(±15pt)。硬编码锚点就是数据列
    x 位置(微信表头词居中且中间 6 列合并, 表头盒 x0 与数据列 x0 不一致);
    先验法对"对方户名跨列碎片"(2026 版 魏建耀@424 / (魏@453 / 商户单号@470)
    也稳定。某列找不到近邻词 → 返回 None(格式变化过大, 回退硬编码)。"""
    dates = [w for w in words if w[1] > header_y and WECHAT_DATE_PATTERN.match(w[4])]
    if not dates:
        return None
    main_ys = sorted({round(d[1], 1) for d in dates})[:max_rows]
    anchors = []
    for px in (c[1] for c in WECHAT_COLS):
        best = None
        for y in main_ys:
            for w in words:
                if (abs(w[1] - y) < 2.0 and (w[2] - w[0]) >= 0.5
                        and abs(w[0] - px) <= 15.0):
                    if best is None or abs(w[0] - px) < abs(best[0] - px):
                        best = w
        if best is None:
            return None
        anchors.append(best[0])
    if any(a is None for a in anchors):
        return None
    return anchors


def _wechat_dynamic_bounds(words, anchors, debug=False):
    """微信动态列锚点(P2-5): 视觉读首页表头 → 生成 8 列 x 锚点。
    注意: 微信 PDF 表头中间 6 列(交易时间~金额(元))常渲染为一个合并词, 且
    表头词居中(交易单号盒 x0=78, 数据列 x0=40), 直接按表头盒反查不可靠。
    流程: 1) 视觉锚点校验 8 列名(顺序/内容); 2) 数据列 x0 锚点恢复
    (_wechat_data_anchors, 与硬编码同源); 3) 与硬编码容差内一致 → 用硬编码
    (保证既有格式逐格一致); 明显偏移(改版) → 用数据锚点动态边界;
    4) 失败 → None 回退硬编码。"""
    if not anchors:
        return None
    hard_xs = [c[1] for c in WECHAT_COLS]
    data_xs = _wechat_data_anchors(words)
    if data_xs and len(data_xs) == len(WECHAT_COLS):
        # 与硬编码容差内一致 → 直接用硬编码(当前格式基线行为)
        if all(abs(x - px) <= 10.0 for x, px in zip(data_xs, hard_xs)):
            if debug:
                log("[DEBUG] 微信动态锚点: 数据锚点与硬编码一致, 沿用硬编码边界", debug)
            return wechat_column_bounds()
        # 改版: 数据锚点明显偏移 → 动态边界
        names = [c[2] for c in wechat_column_bounds()]
        cand = []
        for i, x in enumerate(data_xs):
            lo = 0.0 if i == 0 else (data_xs[i - 1] + x) / 2.0
            hi = 595.0 if i == len(data_xs) - 1 else (x + data_xs[i + 1]) / 2.0
            cand.append((lo, hi, names[i]))
        log(f"[DEBUG] 微信动态锚点启用(改版): {[(round(lo,1), round(hi,1)) for lo,hi,_ in cand]}", debug)
        return cand
    log("[DEBUG] 微信动态锚点恢复失败, 回退硬编码", debug)
    return None


def extract_wechat_statement(doc, debug=False, anchors=None, bounds=None):
    """微信支付交易明细证明专用提取。返回 (表头列表, 数据行列表, meta)
    anchors: 视觉读出的首页表头列名(动态列锚点, 失败回退硬编码);
    bounds: 直接提供的列边界列表(格式描述符 columns, 优先于 anchors)。"""
    if bounds is not None:
        bounds = [(float(lo), float(hi), name) for lo, hi, name in bounds]
    elif anchors:
        words0 = doc[0].get_text("words")
        if doc[0].rotation != 0:
            words0 = normalize_page_words(doc[0], words0)
        bounds = _wechat_dynamic_bounds(words0, anchors, debug=debug)
        if bounds is None:
            bounds = wechat_column_bounds()
    else:
        bounds = wechat_column_bounds()
    header_names = [b[2] for b in bounds]
    # 微信日期固定 YYYY-MM-DD; 不能用默认 4 种日期正则——商户单号的 8 位数字片段
    # (如 20105739) 会被 ^\d{8}$ 误判为日期行, 导致续行被错误切成新记录(+1000 条)
    wechat_date_patterns = [WECHAT_DATE_PATTERN]
    # 交易对方列 x 范围(用于剥离混入的商户单号)
    opp_lo, opp_hi = bounds[6][0], bounds[6][1]
    records = []
    page_meta = []
    for pno in range(len(doc)):
        pg_records = 0
        pg_first_date = pg_last_date = None
        page = doc[pno]
        words = page.get_text("words")
        if page.rotation != 0:
            words = normalize_page_words(page, words)
        date_rows_all = find_date_rows(words, wechat_date_patterns)
        min_date_y = min((w[1] for w in date_rows_all), default=10**9)
        # 仅首页有表头; 后续页数据顶到页面顶部(用本页最小日期行兜底)
        if pno == 0:
            y_min = 214.9 + 3
        else:
            y_min = max(0, min(min_date_y - 6, 214.9 - 90))
        # 页脚行级检测
        # 行级文本预聚合(微信页脚检测用): 一次 round/join, region 过滤复用行键
        page_rows = defaultdict(list)
        word_row = []
        for w in words:
            rk = round(w[1], 1)
            word_row.append(rk)
            page_rows[rk].append(w)
        page_row_text = {y: " ".join(w[4] for w in ws) for y, ws in page_rows.items()}
        footer_ys = [
            y for y in page_rows
            if y > y_min + 100 and is_footer_word(page_row_text[y])
        ]
        page_footer_y = min(footer_ys, default=10**9)
        footer_row_keys = set(footer_ys)
        region = []
        for wi, w in enumerate(words):
            if not (y_min < w[1] < page_footer_y):
                continue
            if w[4].startswith("_") and len(w[4]) > 10:
                continue
            if word_row[wi] in footer_row_keys:
                continue
            if (w[2] - w[0]) < 0.5:  # 零宽隐藏词过滤(与通用分支一致)
                continue
            region.append(w)
        if not region:
            continue
        date_rows = [w for w in date_rows_all if y_min < w[1] < page_footer_y]
        if not date_rows:
            log(f"[DEBUG] 第{pno+1}页未找到日期行, 跳过", debug)
            continue
        for cluster in slice_records(region, date_rows, page_footer_y):
            cells = [[] for _ in bounds]
            for w in cluster:
                x0, x1, text = w[0], w[2], w[4]
                # 交易对方列词若 x 跨越到商户单号列(尾部字母数字串 = 商户单号混入,
                # 如 '笑厨饮品店bUsv1tw5WT0HSP6j' x0=422 x1=555) → 拆词分配
                if opp_lo <= x0 < opp_hi and x1 > opp_hi:
                    m = re.search(r"([A-Za-z0-9_-]{6,})$", text)
                    if m:
                        head, tail = text[:m.start()], m.group(1)
                        if head:
                            cells[6].append(head)
                        cells[7].append(tail)
                        continue
                for ci, (lo, hi, name) in enumerate(bounds):
                    if lo <= x0 < hi:
                        cells[ci].append(text)
                        break
            rec = []
            for ci, (lo, hi, name) in enumerate(bounds):
                # 交易时间列(日期+时间两行)用空格连接, 其余列直接拼接(跨行字段)
                sep = " " if "时间" in name else ""
                val = sep.join(cells[ci]).strip()
                # 归一化 PDF 字体变体字符: 康熙部首"⼊"(U+2F0A) → 标准"入"(U+5165)
                # (微信"收⼊"的"入"在部分 PDF 用部首字形渲染, 会导致收/支分类与表格内容不一致)
                if "\u2f0a" in val:
                    val = val.replace("\u2f0a", "\u5165")
                rec.append(val)
            if any(rec):
                records.append(rec)
                pg_records += 1
                dv = rec[1].split()[0] if rec[1].strip() else None
                if dv:
                    if pg_first_date is None:
                        pg_first_date = dv
                    pg_last_date = dv
        if pg_records:
            page_meta.append({
                "page": pno + 1,
                "records": pg_records,
                "first_date": pg_first_date,
                "last_date": pg_last_date,
                "first_serial": None, "last_serial": None,
                "first_balance": None, "last_balance": None,
                "serial_sequential": False,
            })
    if not records:
        raise RuntimeError("未能提取到任何微信交易记录")
    return header_names, records, {
        "page_count": len(doc), "truncated_words": 0, "wechat": True,
        "pages": page_meta,
    }


def _num(v):
    """把单元格值转 float; None/空/不可解析 → None。
    提取带符号金额(兼容"转账+6,000.00"式 现转标志+金额 合并词, 北京农商银行)。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    # 2026-08-25: 兼容 PDF 无小数位金额词(建行企业明细 '50000.'/'17588.' 等,
    # 源 PDF 渲染省略 .00): 正则从 强制两位小数 放宽为 可选 1-2 位小数。
    m = re.search(r"[+-]?[\d,]+(?:\.\d{1,2})?", s)
    if m is None:
        return None
    try:
        return float(m.group(0).replace(",", ""))
    except (ValueError, TypeError):
        return None


def compute_wechat_stats(header_names, rows):
    """微信流水提取统计: 笔数 + 按收/支/其他列汇总收入/支出合计(金额全为正)。
    2026-08-21: 改为直接基于记录行(不再依赖 openpyxl ws, 与流式写出解耦)。
    返回 (report, stats)。"""
    names = [str(n or "") for n in header_names]
    idx_amt = next((i for i, nm in enumerate(names) if "金额" in nm), None)
    idx_io = next((i for i, nm in enumerate(names) if "收/支" in nm), None)
    income = expense = other = 0.0
    for rec in rows:
        io = rec[idx_io] if idx_io is not None and idx_io < len(rec) else None
        av = _num(rec[idx_amt]) if idx_amt is not None and idx_amt < len(rec) else None
        if av is None:
            continue
        if io == "收入":
            income += av
        elif io == "支出":
            expense += av
        else:
            other += av
    n = len(rows)
    report = [
        f"数据行数: {n}",
        f"收入合计: {income:,.2f}",
        f"支出合计: {expense:,.2f}",
        f"其他(不计收支): {other:,.2f}",
    ]
    stats = {"rows": n, "income": income, "expense": expense, "other": other}
    return report, stats


# ---------- 记录提取 ----------
def find_date_rows(words, patterns, cols=None):
    """找出所有记录起始日期行(匹配日期正则的 word)。返回排序后的列表。
    2026-08-17 扩展:
    - 跳过纯页码词(如 "1/3", PAGE_FOOTER_RE 整行锚定形态), 避免把页脚当日期;
    去重: 同一 y 只保留第一个日期词。
    2026-08-18 扩展(cols 参数): 通用分支传入识别出的列, 非粘合形态的日期词必须
    落在"日期/时间"列, 否则 8 位数字账号/凭证号会被误判为记录锚点(建行 17 列
    企业明细: 对方账号列出现身份证号片段 19200309/20020217, 同样形如合法日期)。
    粘合日期词(GLUDED_DATE_RE, 日期拼进摘要词尾)不受列限制——它本就落在摘要列。"""
    date_words = []
    for w in words:
        text = w[4]
        # 纯页码词("1/3"/"18/18", 多在页面右侧)不是日期; 但 MM/DD 日期("09/01")
        # 也在交易表左列, 不能按形态一概跳过 → 仅跳过右侧(x0>200)的 N/M 词
        if w[0] > 200 and PAGE_FOOTER_RE.match(text):
            continue
        matched = False
        glued = GLUED_DATE_RE.match(text)
        for pat in patterns:
            if pat.match(text):
                matched = True
                break
        # 摘要+日期粘合词(如建行 "支付机构提现20250103"): 日期被 PDF 拼进摘要词尾,
        # 单词整体不匹配日期正则, 但按"尾随 8 位日期"仍应作为记录起始锚点。
        if not matched and glued:
            matched = True
        if matched:
            # 列限定(仅对正则命中的日期词): 防止 8 位账号/凭证号被当作记录锚点
            if cols is not None and not glued:
                ci = match_column(w[0], text, cols)
                if ci is None or not any(k in cols[ci][2] for k in ("日期", "时间")):
                    continue
            date_words.append(w)
    # 去重: 同一 y 只保留第一个日期词
    seen_y = set()
    out = []
    for w in sorted(date_words, key=lambda x: x[1]):
        if round(w[1], 1) not in seen_y:
            seen_y.add(round(w[1], 1))
            out.append(w)
    return out


def extract_statement(doc, date_patterns, debug=False, columns_template=None,
                      layout=None, footer_extra=None, semantics=None,
                      header_anchors=None):
    """主流程: 解析 PDF → 返回 (表头列表, 数据行列表, 元信息dict)。
    接收已打开的 doc(不再二次 open PDF); 每页记录切分走 slice_records 二分。
    columns_template(可选): 视觉锚点 onboarding 生成的列模板
    [(lo, hi, 列名), ...]; 提供时跳过启发式列识别(表头页仍需用于 y 定位)。
    layout(可选): 格式描述符的布局提示, "columnar-multiline" 时直接走 17 列
    签名分支的边界+长文本逻辑(不再要求表头签名/动态形态门控)。
    footer_extra(可选): 格式描述符的页脚样例关键词, 附加到页脚行检测
    (仅在 y > 表头+100 的行上生效, 不会误伤数据区)。
    semantics(可选): 与 cols 对齐的语义标签列表(date/debit/credit/balance/
    currency/counterparty_name/counterparty_account/bank/note/summary/serial/
    type/number/other), 优先于列名字符串用于日期列/长文本列/统计判定。
    header_anchors(可选): 视觉读出的表头列名列表。表头检测失败(视觉兜底)时,
    对每一页用锚点反推表头带(find_header_band_by_anchors), 支持"每页表头
    y 位置不同"的格式(民生/华夏/邮储等)在启发式检测失效后仍能逐页定位表头。"""
    page_count = len(doc)
    if page_count == 0:
        raise RuntimeError("PDF 为空")

    # ---- 水印/固定元素检测 ----
    # 对账单水印(户名/账号)与页脚公告在每页固定 (x,y) 位置重复出现。
    # 注意: 常见摘要词(如"银联入账")也会因行数固定而跨页同位置, 不能整词过滤;
    #       仅在"与数据同槽冲突"或"孤立行槽"时才判定为水印并移除(见 clean_overlaps)
    # 先做旋转归一化: 把所有页 words 转换到视觉坐标系(rect), 与 detect_header/extract 等
    # 后续逻辑(假设 x=列、y=行)保持一致。rotation=0 时 normalize_page_words 直接返回原 words
    page_words_list = []
    for pno in range(page_count):
        pw = doc[pno].get_text("words")
        if doc[pno].rotation != 0:
            pw = normalize_page_words(doc[pno], pw)
        page_words_list.append(pw)
    pos_counter = Counter()
    for pw in page_words_list:
        for w in pw:
            pos_counter[(w[4], round(w[0], 1), round(w[1], 1))] += 1
    # 水印判定阈值: 必须过半页数每页出现(如户名水印每页 8 处全命中)。
    # 阈值太低会把合法数据误判为水印: 3 页文档中 2/3 页同位置出现的词并不罕见
    # (旅立方续行 "公司 司北京顺义支行 回单号" 在 2/3 页同坐标, 若按 2 页阈值会
    # 被 clean_overlaps 当页脚公告删除 → 户名截断)。下限 3 页:
    #  - 3 页短账单的文字水印 3/3 页全命中 → 检出 ✓
    #  - 2/3 页同位置的合法续行/摘要词不被误删 ✓
    min_pages = max(3, int(page_count * 0.5))
    fixed_keys = {k for k, v in pos_counter.items() if v >= min_pages}
    # 网格路径专用 fixed 子集(2026-08-29): 仅"位置唯一(同词x去重y数==1)"且
    # 非日期/时间形态的词视为真固定元素。启发式路径仍用 full fixed_keys(不动),
    # 该子集只用于网格行带路径的 clean_overlaps —— 浦发个人流水每页首条记录
    # 布局一致, 交易账号续段 '82138'(143 页同坐标但 y 位置多变)被 full 判定
    # 当水印删掉 → 账号截断; 网格路径用此子集可保住这类合法数据词。
    fixed_keys_unique = None
    try:
        if fixed_keys:
            _y_pos = {}
            for pw in page_words_list:
                for w in pw:
                    _y_pos.setdefault((w[4], round(w[0], 1)), set()).add(round(w[1], 1))
            fixed_keys_unique = {
                k for k in fixed_keys
                if len(_y_pos.get((k[0], k[1]), ())) <= 1
                and not is_date_like_word(k[0])
                and not is_time_word(k[0])
            }
    except Exception:  # noqa: BLE001
        fixed_keys_unique = fixed_keys
    if fixed_keys:
        log(f"[DEBUG] 检测到 {len(fixed_keys)} 个跨页固定元素候选, 仅清理冲突/孤立项", debug)

    # ---- 扫描各页找第一个含表头的页(首页可能是封面/摘要/条款页, 无交易表) ----
    header_page_idx = None
    header = None
    cols = None
    band_bottom = None
    for pno in range(page_count):
        cand_header = detect_header_line(page_words_list[pno])
        if cand_header is not None:
            cand_cols, cand_band_bottom = detect_column_boundaries(page_words_list[pno], cand_header)
            if cand_cols:
                header_page_idx = pno
                header = cand_header
                cols = columns_template if columns_template is not None else cand_cols
                band_bottom = cand_band_bottom
                header_y = cand_header[0]
                break
        # 视觉兜底: detect_header_line 失败(特殊列名/盖章遮挡)时用锚点反推表头带
        elif columns_template is not None and header_anchors:
            ah = _anchor_page_header(page_words_list[pno], header_anchors, debug)
            if ah is None:
                continue
            h_y, h_band, h_bb = ah
            header_page_idx = pno
            header = (h_y, h_band, "single")
            cols = columns_template
            band_bottom = h_bb
            header_y = h_y
            break
    if header is None or cols is None:
        raise RuntimeError(
            "未能自动识别表头行。请确认这是银行对账单类 PDF，"
            "或使用 --header-keywords 自定义表头关键词。"
        )
    log(f"[DEBUG] 表头页={header_page_idx+1}, 表头行 y={header_y}, 识别列: {[c[2] for c in cols]}", debug)

    # 视觉锚点模板: 用数据词 x 聚类精修边界(锚点中点边界对"表头居中/数据左对齐"
    # 或"表头窄/数据宽"格式会把相邻列数据错分, 与启发式数据聚类修正保持一致)
    if columns_template is not None:
        cols = refine_cols_with_data_x(page_words_list[header_page_idx], cols, band_bottom)
        log(f"[DEBUG] 列模板数据聚类精修: {len(cols)} 列", debug)

    # ---- 完整表格网格检测(2026-08-18) ----
    # 若 PDF 用填充矩形逐格绘制表格(建行企业账户 17 列等), 网格直接给出精确
    # 列边界(优于"表头中点+数据聚类"的启发式近似)与每页行带(一行=一条记录)。
    # 仅当网格列数 == 启发式列数时替换列边界(列名仍沿用启发式识别结果);
    # 行带在每页另做"行带数==记录数"校验, 失败回退 slice_records。
    # 每页 get_drawings 成本高(建行 17 列 116 页实测 ~0.6s): 页面级缓存,
    # 网格路径下每页只取一次(表头页与页面循环共用), 非网格格式零额外开销。
    _drawings_cache = {}

    def _page_drawings(pno):
        d = _drawings_cache.get(pno)
        if d is None:
            d = doc[pno].get_drawings()
            _drawings_cache[pno] = d
        return d

    def _page(pno):
        return doc[pno]

    def _page_words(pno):
        return page_words_list[pno]

    grid_edges = None
    try:
        grid_edges = detect_table_grid(_page_drawings(header_page_idx))
    except Exception:  # noqa: BLE001
        grid_edges = None
    if grid_edges is not None:
        vv, _hh = grid_edges
        if len(vv) - 1 == len(cols):
            cols = [(vv[i], vv[i + 1], cols[i][2]) for i in range(len(vv) - 1)]
            log(f"[DEBUG] 表格网格启用: {len(vv)-1} 列精确边界", debug)
        else:
            grid_edges = None

    # ---- 汇总行与页脚关键词 ----
    # 数据区范围: 表头带之下, 页脚(含"打印时间"等的行)之上。
    # 页脚按**行级**文本检测(页码"第 N 页/共 M 页"常被拆成多个词, 单词级正则会漏匹配)
    hp_rows = defaultdict(list)
    for w in page_words_list[header_page_idx]:
        hp_rows[round(w[1], 1)].append(w)
    footer_y = min(
        (y for y, ws in hp_rows.items()
         if y > band_bottom + 100 and is_footer_word("".join(w[4] for w in ws))),
        default=10**9,
    )

    # 表格右缘(列边界最大 hi): 超出该右缘 +15pt 的行视为页边注记/侧栏说明
    table_right = max(c[1] for c in cols)

    # ---- 提取所有页数据 ----
    header_names = [c[2] for c in cols]
    records = []
    truncated_note = 0
    page_meta = []  # QA 用: 每页记录数/首末日期/首末流水号/首末余额
    # 语义标签(可选): 列名子串判定的替代/优先信号
    sem = semantics if semantics and len(semantics) == len(cols) else None
    def _sem_or_name(ci, *keys):
        """语义优先的列判定: 有语义标签时按标签匹配, 否则回退列名字符串。"""
        if sem is not None:
            return any(sem[ci] == k for k in keys)
        return any(k in cols[ci][2] for k in keys)
    # 日期列: 语义优先(特殊列名如"交易时间"已含中文, 但双语列名/语义更稳)
    date_col_idx = next((i for i in range(len(cols))
                         if sem is not None and sem[i] == "date"), None)
    if date_col_idx is None:
        date_col_idx = next((i for i, (_lo, _hi, nm) in enumerate(cols)
                             if any(k in nm for k in ("日期", "时间"))), None)
    serial_col_idx = next((i for i in range(len(cols))
                           if sem is not None and sem[i] == "serial"), None)
    if serial_col_idx is None:
        # 回退匹配优先真正的流水号/序号列; "凭证号"只是凭证编号,
        # 仅在无更好候选时才采用(民生表头"凭证号码"在"流水号"之前,
        # 旧逻辑会把凭证号当流水号导致跨页连续性误报)。
        serial_col_idx = next((i for i, (_lo, _hi, nm) in enumerate(cols)
                               if "流水号" in nm), None)
        if serial_col_idx is None:
            serial_col_idx = next((i for i, (_lo, _hi, nm) in enumerate(cols)
                                   if "序号" in nm), None)
        if serial_col_idx is None:
            serial_col_idx = next((i for i, (_lo, _hi, nm) in enumerate(cols)
                                   if "凭证号" in nm), None)
    serial_name = cols[serial_col_idx][2] if serial_col_idx is not None else ""
    # 仅"序号"型列(如天津银行 1..1713)才是页间连续的; 随机流水号/凭证号不是
    serial_sequential = bool(re.search(r"序号", serial_name))
    balance_col_idx = next((i for i in range(len(cols))
                            if sem is not None and sem[i] == "balance"), None)
    if balance_col_idx is None:
        balance_col_idx = next((i for i, (_lo, _hi, nm) in enumerate(cols)
                                if "余额" in nm), None)
    debit_col_idx = next((i for i in range(len(cols))
                          if sem is not None and sem[i] == "debit"), None)
    if debit_col_idx is None:
        debit_col_idx = next((i for i, (_lo, _hi, nm) in enumerate(cols)
                              if "借方" in nm), None)
    credit_col_idx = next((i for i in range(len(cols))
                           if sem is not None and sem[i] == "credit"), None)
    if credit_col_idx is None:
        credit_col_idx = next((i for i, (_lo, _hi, nm) in enumerate(cols)
                               if "贷方" in nm), None)
    # 2026-08-22 农业银行账户历史明细: 独立"收入金额/支出金额"两列,
    # 映射到 credit/debit(page meta/QA 跨页衔接使用); 不影响工作簿列顺序。
    income_col_idx = next((i for i, (_lo, _hi, nm) in enumerate(cols)
                           if "收入金额" in nm or nm.strip() == "收入"), None)
    expense_col_idx = next((i for i, (_lo, _hi, nm) in enumerate(cols)
                            if "支出金额" in nm or nm.strip() == "支出"), None)
    if credit_col_idx is None:
        credit_col_idx = income_col_idx
    if debit_col_idx is None:
        debit_col_idx = expense_col_idx
    amount_col_idx = next((i for i in range(len(cols))
                           if sem is not None and sem[i] == "number"), None)
    if amount_col_idx is None:
        amount_col_idx = next((i for i, (_lo, _hi, nm) in enumerate(cols)
                               if ("金额" in nm and "发生" not in nm)
                               or nm.strip() == "发生额"), None)
    if income_col_idx is not None and expense_col_idx is not None:
        amount_col_idx = None
    # 17 列签名与布局提示与页无关, 循环外计算一次(2026-08-21 性能优化: 原每页重复)
    _names17 = [c[2] for c in cols]
    _sig17 = (any("交易时间" in n for n in _names17)
              and any("记账日期" in n for n in _names17)
              and any(("账户明细编号" in n) or ("交易介质编号" in n) for n in _names17))
    force_multiline = layout == "columnar-multiline"
    for pno in range(page_count):
        # 进度回调(PYODIDE/H5): 每页开始时上报 (当前页, 总页数); 非 PYODIDE 为
        # no-op 开销(模块级 PROGRESS_CB 默认 None)。
        if PROGRESS_CB is not None:
            try:
                PROGRESS_CB(pno + 1, page_count)
            except Exception:  # noqa: BLE001 (进度上报失败不影响提取)
                pass
        # 跳过表头页之前的封面/摘要/条款页(无交易表的页)
        if pno < header_page_idx:
            continue
        words = page_words_list[pno]
        # QA 用每页统计(记录数/首末日期/流水号/余额)
        pg_records = 0
        pg_first_date = pg_last_date = None
        pg_first_serial = pg_last_serial = None
        pg_first_balance = pg_last_balance = None
        pg_first_debit = pg_first_credit = None
        pg_first_amount = pg_last_amount = None
        # 每页动态检测本页表头 y(部分银行首页表头在信息区下方, 后续页表头顶到页面顶部)
        page_header = detect_header_line(words)
        fb_bb = None
        if page_header is None and header_anchors:
            ah = _anchor_page_header(words, header_anchors, debug)
            if ah is not None:
                page_header = (ah[0], ah[1], "single")
                fb_bb = ah[2]
        page_header_y = page_header[0] if page_header is not None else header_y
        if page_header is not None:
            header_band = page_header[1]
            header_text = " ".join(w[4] for w in header_band)
            # 表头带下界: 命中完整列名的表头词最大 y(拆词段"间/生额"不计, 无害地留在 region)
            hit_ys = [w[1] for w in header_band if any(n in w[4] for n in SPLIT_NAMES)]
            # 锚点反推带优先与启发式一致: 用 SPLIT_NAMES 命中词最大 y
            # (建行25 带底=89 但 hit_ys max=77, y_min=80 才能让网格首行带内的
            # 表头残词"水号"进入 region, 与启发式基线逐格一致);
            # 特殊列名不在 SPLIT_NAMES 时(视觉兜底初衷)才回退锚点带底 fb_bb。
            page_band_bottom = max(hit_ys, default=fb_bb if fb_bb is not None else page_header_y)
        else:
            header_text = ""
            page_band_bottom = band_bottom
        page_has_header = sum(1 for k in DEFAULT_HEADER_KEYWORDS if k in header_text) >= 2
        # 无表头页数据可能顶到页面顶部(y 远小于首页表头): 以该页最小日期行起点兜底,
        # 避免把顶部第一条记录过滤掉(顺义银座村镇银行顶部记录 y=63, 首页表头 y=162)
        date_rows_all = find_date_rows(words, date_patterns, cols)
        min_date_y = min((w[1] for w in date_rows_all), default=10**9)
        if page_has_header:
            y_min = page_band_bottom + 3
        else:
            y_min = max(0, min(min_date_y - 6, header_y - 90)) if min_date_y != 10**9 else max(0, header_y - 90)
        # 行级文本预聚合(2026-08-21 性能优化): 一次 round + 一次 join 建行级字典,
        # 页脚/汇总/次级表头/region 过滤全部复用, 消除每行多次 join 与重复取整
        # (建行 17 列 116 页实测 str.join ~0.47s → 单次聚合)。
        page_rows = defaultdict(list)
        word_row = []
        for w in words:
            rk = round(w[1], 1)
            word_row.append(rk)
            page_rows[rk].append(w)
        page_row_text = {y: " ".join(w[4] for w in ws) for y, ws in page_rows.items()}
        row_has_date = {
            y: any(any(p.match(w[4]) for p in date_patterns) for w in page_rows[y])
            for y in page_rows
        }
        # 每页动态识别页脚行(页码"第 N 页/共 M 页"/"温馨提示"等; 行级文本正则匹配,
        # 既避免单词级漏匹配拆分的页码词, 也避免"共"字误伤数据如"北京公共交通…")
        footer_ys = [
            y for y in page_rows
            if y > page_header_y + 100 and (
                is_footer_word(page_row_text[y])
                or (footer_extra and any(fw and fw in page_row_text[y] for fw in footer_extra))
            )
        ]
        page_footer_y = min(footer_ys, default=footer_y)
        footer_row_keys = set(footer_ys)
        # 汇总行整行排除("借方发生总额/合计笔数/期末余额" 等; 无日期才判为汇总行,
        # 避免误删"摘要含'合计'字样"的数据行; 避免"合计"行数字残留进记录)
        summary_rows = {
            y for y in page_rows
            if BAL_SUMMARY_RE.search(page_row_text[y])
            or (SUMMARY_RE.search(page_row_text[y]) and not row_has_date[y])
        }
        # 次级表头行: 一页内含两段交易表时, 下方段落的表头行会再次命中表头关键词,
        # 作为数据会污染上方末条记录 → 整行排除
        extra_header_ys = set()
        if page_has_header:
            for y in page_rows:
                if y > y_min + 5 and sum(
                        1 for k in DEFAULT_HEADER_KEYWORDS if k in page_row_text[y]) >= 3:
                    extra_header_ys.add(y)
        # 尾部页脚兜底: 最后一条记录之后、无日期的大段页脚文字(银行地址/声明/序列号/
        # 服务费说明等), 取"最后一条**有效**日期行(已排除汇总行) + max(15, 2×日期间距
        # 中位数)"之后的第一个无日期行为页脚下界。
        # 注意: 日期行必须先排除汇总行——否则尾页正文里的日期会把尾部间隔撑大,
        # 让整段页脚文字落进最后一条记录。
        date_rows_all_in_page = [
            w for w in date_rows_all
            if y_min < w[1] < page_footer_y and round(w[1], 1) not in summary_rows
        ]
        if date_rows_all_in_page:
            last_dy = date_rows_all_in_page[-1][1]
            if len(date_rows_all_in_page) >= 2:
                gaps2 = sorted(date_rows_all_in_page[i + 1][1] - date_rows_all_in_page[i][1]
                               for i in range(len(date_rows_all_in_page) - 1))
                tail_margin = max(15.0, 2.0 * gaps2[len(gaps2) // 2])
            else:
                tail_margin = 45.0
# 网格优先(2026-08-29): 网格行带路径可用(与主流程网格分支同一门控:
            # grid_edges 非空, 已含"网格列数==启发式列数"校验)且本页网格行带数 ==
            # 日期行数时, 记录边界由网格行带精确给出, 页脚天然隔离在网格之外 ——
            # 跳过尾部页脚兜底收紧。否则启发式 tail_cut(末条日期行+45pt)会把末页
            # 长续行(浦发个人流水末页摘要 ID623336334 延伸到 y=126, 而
            # tail_cut=91)误判为页脚整段剔除。
            # 可用性判据必须完整: 浦发企业回单网格 17 列 != 启发式 9 列, 行带路径
            # 实际不可用(grid_edges=None), 若误判可用而跳过 tail, 每页末条记录会
            # 吞进页尾公告。
            _grid_hh2 = None
            if grid_edges is not None:
                try:
                    _gv2, _gh2 = detect_table_grid(_page_drawings(pno))
                    if _gv2 is not None:
                        _grid_hh2 = _gh2
                except Exception:  # noqa: BLE001
                    _grid_hh2 = None
            _grid_path_ok = (
                _grid_hh2 is not None
                and len(_grid_hh2) - 1 == len(date_rows_all_in_page)
            )
            if not _grid_path_ok:
                tail_cut = last_dy + tail_margin
                tail_footer = min(
                    (y for y in page_rows if y > tail_cut and not row_has_date[y]),
                    default=10**9,
                )
                page_footer_y = min(page_footer_y, tail_footer)
            # 边注行兜底: 最后日期行之后、tail_cut 之前, 无日期且超出表格右缘的行
            # → 提前截断, 防止整段侧栏文字混入末条记录; 正常续行都在表格宽度内。
            marginalia = min(
                (y for y in page_rows
                 if last_dy + 8 < y < tail_cut
                 and not row_has_date[y]
                 and any(w[0] > table_right + 15 for w in page_rows[y])),
                default=10**9,
            )
            page_footer_y = min(page_footer_y, marginalia)
        # 过滤: 数据区 + 非横线 + 非汇总。
        # 双语表头英文镜像词(浦发 'Date'/'Debit'/'Counterparty Institution' 等,
        # 位于中文表头下方 13-16pt, 无表头关键词命中, 落在 y_min 与首条日期行
        # 之间; 精确边界路径下会被吸入首条记录)整行排除: 首条日期行**上方 ≥16pt**
        # 的纯非中文词不是数据(数据行必含中文或日期锚点)。16pt 阈值是关键:
        # 记录头部段(如建行 17 列账号列首段 '11050138'/'2619-' 在首日期行上方
        # 仅 4-8pt, 属于首条记录行带)必须保留, 否则账号/对方账号/账户明细编号
        # 丢前缀(实测 473 行样本丢 132 格); 浦发/华夏英文镜像词在首日期行上方
        # 19-32pt, 全部 ≥16pt 被排除。注意用已过滤的 date_rows_all_in_page(排除
        # 汇总行/元信息), 不能直接用 date_rows_all —— 华夏"查询起止日期
        # 20260101至…"的词 '20260101' 会命中年份正则, 把首日期行误当成 y=99,
        # 过滤失效。
        _first_date_y = min((w[1] for w in date_rows_all_in_page), default=10**9)
        # 数据行带(词数 >=4 的行): 非数据行的固定/印章词剔除(见 STAMP_CODE_RE)。
        # ICBC 借记明细对角文字印章: 固定词(中国工商银行/12:09:50/2026-04-29/
        # 53221649, 每页同位置→fixed_keys)与每页不同的防伪码(FFFEB5EC2026…)会插在
        # 记录日期行与数据行之间, 把记录切成两个行槽 → 数据行被 clean_overlaps 误判
        # 为"非数据行槽"而整槽删除常量列(账号/储种/序号/币种/钞汇/地区)。先在此剔除,
        # 使日期行与数据行同在数据行槽, 常量列按单列单词保留。
        _ybucket = Counter(round(w[1]) for w in words)
        _data_row_ys = {_y for _y, _c in _ybucket.items() if _c >= 4}
        region = []
        for wi, w in enumerate(words):
            if not (y_min < w[1] < page_footer_y):
                continue
            if w[1] < _first_date_y - 16 and not re.search(r"[\u4e00-\u9fff]", w[4]):
                continue  # 首条日期行上方 ≥16pt 的纯非中文词(英文镜像表头等)
            if w[4].startswith("_") and len(w[4]) > 10:
                continue
            if SUMMARY_RE.search(w[4]):
                continue
            rk = word_row[wi]
            # 页脚/汇总/次级表头行整行排除(按行标识而非仅 y 比较): 页脚词 y0 可能
            # 比取整后的页脚 y 小 0.04pt, 仅用 y < page_footer_y 会让页脚词混入
            # region, 页尾修复后会被并入末条记录(华夏末条对方户名出现 "总2页第1页")。
            if rk in footer_row_keys or rk in summary_rows or rk in extra_header_ys:
                continue
            if META_ROW_RE.search(page_row_text.get(rk, "")):
                continue
            if (w[2] - w[0]) < 0.5:  # 零宽隐藏词(PDF 隐藏元素, 如兴业 'elect_sign' 宽 0.4pt)
                continue
            # 对角印章/防伪码: 无论是否落在数据行, 只要形态/字号命中即剔除。
            # 纯数字 12 位码由 STAMP_WORD_H 联合判定, 不会误删正常账号/流水。
            if _is_stamp_code_word(w):
                continue
            # 跨页固定元素仍限非数据行剔除, 避免误删同位置合法数据。
            # 2026-08-29: 网格路径可用(grid_edges 非空)时用 fixed_keys_unique
            # (位置唯一子集)判定 —— 浦发个人流水每页首条记录续行(账号段 82138 等)
            # 在 fixed_keys 里但 y 位置多变, 非数据行(词数<4)判定会把它误剔 →
            # 账号截断; unique 子集只认真水印。非网格格式(工行宫格水印等)保持
            # full fixed_keys 原行为, 避免水印残留回归。
            _fk_here = fixed_keys_unique if (
                fixed_keys_unique is not None and grid_edges is not None
            ) else fixed_keys
            if (round(w[1]) not in _data_row_ys) and (
                (w[4], round(w[0], 1), round(w[1], 1)) in _fk_here
            ):
                continue
            region.append(w)
        if not region:
            continue
        date_rows = [w for w in date_rows_all_in_page if w[1] < page_footer_y]
        # 合并同记录的多日期列(建行活期明细"交易时间"+"记账日期"两列日期 y 差 <10pt,
        # 若不合并会把同一条记录切成 2 条)。
        # 2026-08-17 修正: 仅当两个日期词落在**不同列**时才合并——同一列内 y 差 <10pt
        # 的两个日期是相邻两行不同记录(行距小的账单若合并会把两笔交易挤进一行)。
        if len(date_rows) > 1:
            merged = [date_rows[0]]
            for w in date_rows[1:]:
                if w[1] - merged[-1][1] < 10:
                    ci_cur = match_column(w[0], w[4], cols)
                    ci_prev = match_column(merged[-1][0], merged[-1][4], cols)
                    if ci_cur is not None and ci_cur != ci_prev:
                        continue  # 同一记录的两个日期列 → 合并为一条
                merged.append(w)
            date_rows = merged
        if not date_rows:
            log(f"[DEBUG] 第{pno+1}页未找到日期行, 跳过", debug)
            continue
        # 记录边界吸收量: 大间距银行(日期行间距中位数 ≥30pt, 如建行活期明细记录
        # 多行段跨日期 ±20pt、邮储摘要跨行)加大吸收; 普通银行(间距 12-20pt)保持 6pt
        # 2026-08-22 农业银行账户历史明细: 行距中位数虽然 ≥30pt, 但交易时间列
        # 是"日期行 + 12pt 下方独立 HH:MM:SS"的两行单元格; 若用 absorb=18,
        # 时间词会被切进下一条记录(日期只剩 00:00、下一条时间列拼接两段时间)。
        # 检测到时间列内、最近日期行下方 ≤20pt 的纯时间词时强制回 6pt。
        date_ys = [d[1] for d in date_rows]
        stacked_time = False
        if date_ys:
            for w in region:
                if not TIME_WORD_RE.match(w[4]):
                    continue
                ci = match_column(w[0], w[4], cols)
                if ci is None or "时间" not in cols[ci][2]:
                    continue
                j = bisect.bisect_right(date_ys, w[1]) - 1
                if j >= 0 and 0 <= w[1] - date_ys[j] <= 20:
                    stacked_time = True
                    break
        if stacked_time:
            absorb = 6
        elif len(date_rows) >= 3:
            gaps = sorted(date_rows[i + 1][1] - date_rows[i][1] for i in range(len(date_rows) - 1))
            d_median = gaps[len(gaps) // 2]
            # 2026-08-18 修正: 上界从 22 收紧到 18 —— 建行企业账户 17 列明细的
            # 多行单元格下段(对方户名/开户机构末行、长备注续行)可延伸到日期下方
            # 20-24pt(个别税务记录达 +88pt, 但对应间距 108pt), 吸收量过大会把
            # 下段切给下一条记录(对方户名截断、上一条尾巴拼进下一条)。
            # 实测该格式: 上延 ≤16pt、下延 ≤24pt、间距最小 36pt(36-24=12 < 18
            # 仍安全); 吸收 18 同时满足"≥上延"与"< 间距-下延"。
            # 普通银行(间距 12-20pt)不受影响(仍走 6pt 分支)。
            absorb = min(18, d_median // 2 + 2) if d_median >= 30 else 6
        else:
            absorb = 6
        # 列式多行单元格布局(建行企业账户 17 列明细): 固定吸收量无法同时覆盖记录
        # 上/下段(上延 20pt、下延 24pt、税务记录备注续行 +88pt), 按列内运行段
        # 逐对计算精确边界; 计算失败时回退吸收法。
        # 门控: 表头签名(同时含"交易时间"+"记账日期"两个日期列 + "账户明细编号/
        # 交易介质编号")——该签名只出现在建行企业账户 17 列明细及其同类变体。
        # 不能仅靠动态 y 形态判断(华夏银行交易机构段在日期上方 2 层、旅立方
        # 上一条记录续行也会形成上方词, 会误触发并把英文表头/上一条续行扫入)。
        # 居中多行单元格(2026-08-21, 浦发电子对账单): 单元格以记录行带垂直居中,
        # 首行可高出日期行 11-17pt, absorb=6 切片会把首行切给上一条记录; 页面
        # 出现 ≥2 个"切片归属 ≠ 最近日期行归属"的非日期词 → 走精确边界(运行段
        # 归属 + 相邻记录中点切分, compute_record_boundaries)。实测该计数在普通
        # 银行(民生/旅立方/北京银行/hqmx/兴业/金小九)每页恒为 0, 不误触发。
        _centered = (not force_multiline and not _sig17
                     and _centered_mis_slice(region, date_rows, cols) >= 2)
        boundaries = None
        excluded_cols = set()
        longtext_cells = {}
        if (force_multiline or _sig17 or _centered) and (
                force_multiline or _centered or is_multiline_column_layout(region, date_rows, cols)):
            boundaries, excluded_cols = compute_record_boundaries(region, date_rows, cols, debug)
            longtext_cells = compute_longtext_cells(region, date_rows, cols, excluded_cols, debug)
        # 重建成功的长文本列: 该列所有记录统一用运行段归属结果(缺失即空), 不再
        # 回退 y 切片词 —— y 切片中的该列词可能是上一条记录的长文本(如空备注
        # 记录的行带里混着上一条记录的备注续行), 回退必然串行。
        rebuilt_cols = set(longtext_cells.keys())

        # 表格网格行带: 网格行 = 记录行, 直接按行带取词(词中心落入行带)。
        # 注意: 长备注/企业流水号单元格仍可纵向跨入下一条记录的行带, 因此
        # 长文本列重建(longtext_cells)在网格路径下同样生效。
        grid_clusters = None
        if grid_edges is not None:
            try:
                _gv, _gh = detect_table_grid(_page_drawings(pno))
            except Exception:  # noqa: BLE001
                _gv, _gh = None, None
            if _gv is not None:
# 表头带判别升级(2026-08-29): 用行带内文本含表头关键词判定,
                # 不再依赖 header_y 落点 —— 仅首页有表头的格式(浦发企业回单/
                # 浦发个人流水), 首页表头 y 会落进后续页某条数据行带, 按 y 匹配
                # 会把该带误当表头带剔除 → 行带数 != 记录数 → 网格路径失效。
                rows = grid_row_bands(_gh, page_header_y, page_footer_y,
                                      words=_page_words(pno),
                                      header_keywords=DEFAULT_HEADER_KEYWORDS)
                if len(rows) == len(date_rows):
                    # 网格路径列归属升级(2026-08-29): 跨列词按字符级坐标切分。
                    # 词级 match_column 只看词起点 x, "王毛敖海621700001"这类
                    # 跨姓名/账号两列的粘连词会整词归入起点列; 网格既然给出
                    # 精确列边界, 就用 rawdict 字符 bbox 把跨列词切开归位。
                    region = split_cross_column_words(_page(pno), region, cols)
                    grid_clusters = [
                        [w for w in region if r0 <= (w[1] + w[3]) / 2 < r1]
                        for r0, r1 in rows
                    ]
                    grid_clusters = [c for c in grid_clusters if c]
        if grid_clusters is not None:
            cluster_iter = enumerate(grid_clusters)
        else:
            cluster_iter = enumerate(slice_records(region, date_rows, page_footer_y,
                                                   absorb=absorb, boundaries=boundaries))
        for ci_cluster, cluster in cluster_iter:
            # 清理水印/页脚公告等固定元素
            # 网格路径(行带精确分界)用"位置唯一"子集 fixed_keys_unique ——
            # 每页首条记录同坐标重复的合法数据词(浦发账号续段)不会被误删;
            # 启发式路径(slice_records)用 full fixed_keys(历史行为不变)。
            _fk = fixed_keys_unique if grid_clusters is not None else fixed_keys
            cluster = clean_overlaps(cluster, _fk, cols, date_patterns, debug)
            if not cluster:
                continue
            group_date = None
            cells = [[] for _ in cols]
            # 2026-08-21 修正: 粘合日期拆分只在该记录"没有独立日期词"时执行。
            # 浦发电子对账单备注 "招行O2O特惠20250923" 中的 8 位数字是业务文本
            # (促销/商户编号), 记录本身已有独立交易日期 2025/09/23 → 不拆分,
            # 避免日期列拼成 "20250923 2025/09/23" 文本、备注丢失 "20250923"。
            # 建行 "支付机构提现20250103" 等无独立日期词的记录仍正常拆分。
            has_date_word = (date_col_idx is not None and any(
                match_column(w[0], w[4], cols) == date_col_idx
                and any(p.match(w[4]) for p in date_patterns)
                for w in cluster))
            for w in cluster:
                ci = match_column(w[0], w[4], cols)
                if ci is not None:
                    # 长文本列(备注/企业流水号)单元格由 compute_longtext_cells 按
                    # 列内运行段独立重建(长文本纵向与下一条记录重叠, y 切片会
                    # 截断/串行); 有重建结果时跳过 y 切片中的长文本词, 避免重复。
                    if ci in rebuilt_cols:
                        continue
                    # 摘要+日期粘合词(如建行 "支付机构提现20250103"): 拆出头文本与
                    # 尾随日期, 日期归入交易日期列, 避免摘要格残留 "支付机构提现20250103"
                    m = GLUED_DATE_RE.match(w[4])
                    if m and date_col_idx is not None and ci != date_col_idx and not has_date_word:
                        head, d = m.group(1), m.group(2)
                        if head:
                            cells[ci].append((w[1], head))
                        cells[date_col_idx].append((w[1], d))
                    else:
                        # (y, text): 列内按 y 排序拼接, 修正跨行字段词序错乱
                        cells[ci].append((w[1], w[4]))
            for ci in rebuilt_cols:
                cells[ci] = sorted(longtext_cells[ci].get(ci_cluster, []), key=lambda p: p[0])
            # 金额列内的时间形态词重分类到摘要/用途列(民生个人版摘要嵌入时间被拆词,
            # x 落入"交易金额"列 → '0554023,817.46' 脏值)。无摘要/用途列则丢弃该时间词。
            for ci, (lo, hi, name) in enumerate(cols):
                if "金额" in name and "余额" not in name:
                    keep = []
                    for item in cells[ci]:
                        y_v, v = item
                        if is_time_word(v):
                            moved = False
                            for j, (_, _, n2) in enumerate(cols):
                                if "摘要" in n2 or "用途" in n2:
                                    cells[j].append((y_v, (" " if cells[j] else "") + v))
                                    moved = True
                                    break
                            if not moved:
                                continue
                        else:
                            keep.append(item)
                    cells[ci] = keep
            # 摘要/用途列首词为纯时间词(HH:MM:SS) → 归入"时间"类列。
            # (民生个人版表头"交易时间"仅对齐日期词 x=97, 数据时间词 x=136 被表头
            # 中点边界(135.5)切到摘要列 → 摘要列出现 "06:06:02银联入账…"。)
            for ci, (lo, hi, name) in enumerate(cols):
                if ("摘要" in name or "用途" in name) and cells[ci]:
                    first = cells[ci][0][1].strip()
                    if TIME_WORD_RE.match(first):
                        for j, (_, _, n2) in enumerate(cols):
                            if "时间" in n2:
                                cells[j].append(cells[ci][0])
                                cells[ci] = cells[ci][1:]
                                break
            # 金额/余额类列脏值清理: 页码词 "1/3" 误入金额列 → 丢弃;
            # 多余 "0.00" 续行占位 → 有多余非零金额时丢弃 0.00, 避免金额格拼坏。
            for ci, (lo, hi, name) in enumerate(cols):
                if any(k in name for k in ("金额", "余额", "发生额", "支出", "收入")):
                    kept = []
                    for item in cells[ci]:
                        v = item[1].strip()
                        if re.fullmatch(r"\d{1,3}/\d{1,3}", v):
                            continue  # 页码词
                        kept.append(item)
                    vals = [it[1].strip() for it in kept]
                    if len(vals) > 1 and any(_num(v) not in (None, 0.0) for v in vals):
                        kept = [it for it in kept if not (it[1].strip() in ("0", "0.00"))]
                    cells[ci] = kept
            rec = []
            for ci, (lo, hi, name) in enumerate(cols):
                # 时间类列(日期+时分秒分两行)用空格连接; 对手信息列含 % 机构代码词时也用空格分隔
                sep = " " if ("时间" in name or "日期" in name
                              or ("对手" in name and any(v[1].startswith("%") for v in cells[ci]))) else ""
                # 列内按 y(行)排序拼接: 修正 PDF 跨行字段词序错乱(如华夏银行交易机构
                # 跨 2-3 行时 pymupdf 词序与视觉行序不一致)
                val = sep.join(t for _, t in sorted(cells[ci], key=lambda p: p[0])).strip() if cells[ci] else ""
                # 用途/摘要类列若含对手机构代码(%1000050201%02%99%%000 或其截断 %15302...),
                # 剥离 % 及其后内容(对手信息列已有完整账号/户名; PDF 渲染常把 % 词并入用途行)
                if "对手" not in name and "对方" not in name:
                    m = re.search(r"(%\d[\d%]*%+\d*$)|(%\d[\d%]*\.\.\.$)", val)
                    if m:
                        cut = val.find(m.group())
                        if cut > 0:
                            val = val[:cut]
                            truncated_note += 1
                rec.append(val)
            # 丢弃仅含日期(其它列全空)的记录: 汇总行被 SUMMARY 过滤后常剩孤日期词
            non_empty = [v for v in rec if v.strip()]
            if not non_empty:
                continue
            if len(non_empty) == 1 and any(p.match(non_empty[0]) for p in date_patterns):
                continue
            # 组内后续记录日期为空时复制首条记录日期(日期单元格跨行场景)
            if date_col_idx is not None:
                dv = rec[date_col_idx].strip()
                if dv and group_date is None:
                    group_date = dv
                elif group_date is not None and not dv:
                    rec[date_col_idx] = group_date
            records.append(rec)
            pg_records += 1
            if date_col_idx is not None:
                dv = rec[date_col_idx].strip()
                if pg_first_date is None:
                    pg_first_date = dv
                pg_last_date = dv or pg_last_date
            if serial_col_idx is not None:
                sv = rec[serial_col_idx].strip()
                if sv:
                    if pg_first_serial is None:
                        pg_first_serial = sv
                    pg_last_serial = sv
            if balance_col_idx is not None:
                bv = rec[balance_col_idx].strip()
                if bv:
                    if pg_first_balance is None:
                        pg_first_balance = bv
                    pg_last_balance = bv
            # 首笔记录的借方/贷方/单金额必须取第一条记录本身(即使为 0/空),
            # 不能取本页后出现的第一个非空值——否则跨页余额衔接会大量误报。
            if pg_records == 1:
                pg_first_debit = rec[debit_col_idx].strip() if debit_col_idx is not None else None
                pg_first_credit = rec[credit_col_idx].strip() if credit_col_idx is not None else None
                pg_first_amount = rec[amount_col_idx].strip() if amount_col_idx is not None else None
            if amount_col_idx is not None:
                av = rec[amount_col_idx].strip()
                if av:
                    pg_last_amount = av
        if pg_records:
            page_meta.append({
                "page": pno + 1,
                "records": pg_records,
                "first_date": pg_first_date,
                "last_date": pg_last_date,
                "first_serial": pg_first_serial,
                "last_serial": pg_last_serial,
                "first_balance": pg_first_balance,
                "last_balance": pg_last_balance,
                "first_debit": pg_first_debit,
                "first_credit": pg_first_credit,
                "first_amount": pg_first_amount,
                "last_amount": pg_last_amount,
                "serial_sequential": serial_sequential,
            })

    if not records:
        raise RuntimeError("未能提取到任何交易记录，请检查 PDF 内容或日期格式(--date-pattern)")

    return header_names, records, {
        "page_count": page_count,
        "truncated_words": truncated_note,
        "pages": page_meta,
        "wechat": False,
    }


def clean_overlaps(cluster, fixed_keys, cols, date_patterns, debug=False):
    """清理记录内的水印/页脚公告等固定元素。
    判定规则(避免误删"银联入账"等跨页同位置出现的合法摘要):
    1. 冲突规则: 同一行槽(y 容差 2)内同一列出现多个词时, 移除属于固定元素的词
       (典型: 户名水印与日期/摘要词同行重叠)
    2. 孤立规则: 行槽内仅有 1 个词且为固定元素 → 移除
       (典型: 页脚"业务验证"公告, 数据行槽之外的孤立文字)
    """
    # 先全局剔除印章/防伪码词, 避免其作为大字号跨行词污染列内拼接。
    cluster = [w for w in cluster if not _is_stamp_code_word(w)]
    slots = []  # [(y, [words])]
    # 行槽容差 4pt: 水印常紧贴数据行(y 差<2), 摘要跨行第一段与水印需并入同一槽以便整槽判定
    for w in sorted(cluster, key=lambda x: (x[1], x[0])):
        if slots and abs(w[1] - slots[-1][0]) <= 4:
            slots[-1][1].append(w)
        else:
            slots.append((w[1], [w]))

    removed = 0
    keep = []
    for y, ws in slots:
        if len(ws) == 1:
            w = ws[0]
            if (w[4], round(w[0], 1), round(w[1], 1)) in fixed_keys and not MASK_WORD_RE.match(w[4]):
                removed += 1          # 孤立固定词(页脚公告等)
                continue
            keep.append(w)
            continue
        # 非数据行槽(行槽内无日期词): 视为水印/页脚/摘要跨行区,
        # 移除其中的固定词, 保留其余(如摘要跨行文字)
        # 数据行槽判定: 独立日期词或摘要+日期粘合词(如建行 "支付机构提现20250103")
        # 都算日期行 —— 否则粘合词记录所在行会被误判为"非数据行槽",
        # 把跨页同位置出现的合法占位词(建行 "***")当固定元素删掉。
        has_date = any(any(p.match(w[4]) for p in date_patterns) or GLUED_DATE_RE.match(w[4]) for w in ws)
        if not has_date:
            for w in ws:
                if (w[4], round(w[0], 1), round(w[1], 1)) in fixed_keys and not MASK_WORD_RE.match(w[4]):
                    removed += 1
                else:
                    keep.append(w)
            continue
        # 数据行槽(含日期词): 同一列内出现多个词时, 移除其中的固定词
        # 例: 户名水印 x=108.2 与日期词 x=112.7 同在"交易日期"列
        by_col = {}
        for w in ws:
            ci = match_column(w[0], w[4], cols)
            by_col.setdefault(ci, []).append(w)
        for ci, col_words in by_col.items():
            if len(col_words) == 1:
                keep.extend(col_words)
                continue
            for w in col_words:
                if (w[4], round(w[0], 1), round(w[1], 1)) in fixed_keys and not MASK_WORD_RE.match(w[4]):
                    removed += 1      # 同列重叠的固定词(户名水印)
                else:
                    keep.append(w)
    if removed and debug:
        log(f"[DEBUG] 清理固定元素 {removed} 词", debug)
    return keep


def match_column(x, text, cols):
    """返回词所属列下标; 处理金额溢出(凭证号码列中出现金额 → 归入金额列);
    对手机构代码词(%...% 形态)重分类到对手/对方列(村镇银行流水该词 x 常与用途列重叠)"""
    for i, (lo, hi, name) in enumerate(cols):
        # 2026-08-22 农业银行账户历史明细: 首列日期词可越出左边界约 0.4pt, 留 1pt 容差
        if lo - 1.0 <= x < hi:
            # 对手机构代码词(%1000050201%02%99%%000 / 截断 %1...) → 归入对手/对方列
            if COUNTERPART_CODE_RE.match(text) or TRUNCATED_CODE_RE.match(text):
                for j, (_, _, n2) in enumerate(cols):
                    if "对手" in n2 or "对方" in n2:
                        return j
                return i
            # 凭证号码列(非流水号)中出现金额格式 → 归入首个金额列
            if ("凭证" in name and "流水" not in name) and AMOUNT_RE.match(text):
                for j, (_, _, n2) in enumerate(cols):
                    if "发生额" in n2 or "金额" in n2 or "余额" in n2:
                        return j
                return i
            return i
    # 右缘越界回退(2026-08-18 建行全部交易明细): 列右边界按表头词宽计算,
    # 长对方账号/户名文本会视觉溢出列宽(如 'AppStore_AppleMusic' 拆词后
    # 'Store' x0=537.4 > 列右缘 532.69, '_' x0=555.2)。起点距最右列右缘
    # ≤30pt 的词仍归入最右列, 避免丢失; 侧栏/边注距离更大, 不会被吸入。
    if cols and x >= cols[-1][1] and x - cols[-1][1] <= 30:
        return len(cols) - 1
    return None


# ---------- 提取统计(供交付前人工比对) ----------
# 列名语义标签(P2-6): 已知银行列名精确映射 + 关键词规则兜底。
# 语义集合: date/debit/credit/balance/currency/counterparty_name/
# counterparty_account/bank/note/summary/serial/type/number/other。
SEMANTIC_EXACT = {
    "交易时间": "date", "交易日期": "date", "记账日期": "date", "日期": "date",
    "借方发生额": "debit", "借方金额": "debit", "借方": "debit",
    "贷方发生额": "credit", "贷方金额": "credit", "贷方": "credit",
    "账户余额": "balance", "余额": "balance", "交易余额": "balance",
    "币种": "currency", "币别": "currency", "钞汇": "currency",
    "对方户名": "counterparty_name", "对方姓名": "counterparty_name",
    "交易对方": "counterparty_name", "对方名称": "counterparty_name",
    "本方户名": "counterparty_name",
    "对方账号": "counterparty_account", "对方卡/账号": "counterparty_account",
    "对方卡号": "counterparty_account",
    "对方行名": "bank", "开户行": "bank", "对方开户行": "bank",
    "对方开户机构": "bank", "对方银行名称": "bank", "交易机构": "bank",
    "交易渠道": "bank",
    "备注": "note", "附言": "note", "交易地点/附言": "note",
    "摘要": "summary", "用途": "summary",
    "流水号": "serial", "凭证号": "serial", "凭证号码": "serial",
    "凭证种类": "serial", "凭证类型": "serial", "序号": "serial",
    "交易单号": "serial", "商户单号": "serial",
    "账户明细编号-交易流水号": "serial", "企业流水号": "serial",
    "交易介质编号": "serial", "外部系统流水": "serial",
    "交易类型": "type", "收/支/其他": "type", "现转标志": "type",
    "收支方向": "type",
    "金额": "number",
    "金额(元)": "number", "发生额": "number", "交易金额": "number",
}
SEMANTIC_RULES = [
    ("date", ["交易时间", "交易日期", "记账日期", "日期"]),
    ("debit", ["借方发生额", "借方金额"]),
    ("credit", ["贷方发生额", "贷方金额"]),
    ("balance", ["余额"]),
    ("counterparty_account", ["对方账号", "对方卡"]),
    ("counterparty_name", ["对方户名", "对方姓名", "交易对方", "本方户名", "对方名称"]),
    ("bank", ["开户", "行名", "交易机构", "交易渠道", "银行名称"]),
    ("currency", ["币种", "币别", "钞汇"]),
    ("serial", ["流水号", "凭证号", "凭证号码", "凭证种类", "凭证类型",
                "交易单号", "商户单号", "介质编号", "序号"]),
    ("type", ["收/支", "交易类型", "现转", "收支"]),
    ("note", ["备注", "附言"]),
    ("summary", ["摘要", "用途"]),
    ("number", ["金额", "发生额"]),
]


def infer_semantic(name):
    """列名 → 语义标签(已知列名精确映射优先, 关键词规则兜底)。"""
    n = str(name).strip()
    if n in SEMANTIC_EXACT:
        return SEMANTIC_EXACT[n]
    for sem, keys in SEMANTIC_RULES:
        if any(k in n for k in keys):
            return sem
    return "other"


def compute_stats(header_names, rows, semantics=None):
    """通用分支提取统计: 笔数 + 借贷方合计(分列格式) 或 收入/支出合计(单金额列)。
    2026-08-21: 改为直接基于记录行(不再依赖 openpyxl ws, 与流式写出解耦)。
    返回 (report, stats)。不做余额连续性等自动校验——数据质量由代理按 SKILL.md
    「交付前抽样比对」人工核对(2026-08-18 起, 校验逻辑不再比提取更复杂)。"""
    report = []
    n = len(rows)
    names = [str(nm or "") for nm in header_names]
    # 2026-08-23 统一语义判定: 始终通过 infer_semantic 获取每列语义标签,
    # 外部 semantics 参数(视觉 onboarding)可覆盖自动推断。
    auto_sem = [infer_semantic(nm) for nm in names]
    sem = (semantics if semantics and len(semantics) == len(names) else auto_sem)
    if sem is not None:
        idx_debit = next((i for i, s in enumerate(sem) if s == "debit"), None)
        idx_credit = next((i for i, s in enumerate(sem) if s == "credit"), None)
        idx_amount = (next((i for i, s in enumerate(sem) if s == "number"), None)
                      if idx_debit is None and idx_credit is None else None)
    # 2026-08-22 农业银行账户历史明细: 表头为"收入金额/支出金额"两列,
    # 应分别统计; 之前只识别第一个含"金额"列, 导致支出合计恒为 0。
    idx_income = next((i for i, nm in enumerate(names)
                       if "收入金额" in nm or nm.strip() == "收入"), None)
    idx_expense = next((i for i, nm in enumerate(names)
                        if "支出金额" in nm or nm.strip() == "支出"), None)
    if idx_income is not None and idx_expense is not None:
        idx_amount = None
    # 2026-08-23 北京银行等格式: 单"发生额"列全为正值, 无借贷分列、无收/支方向列。
    # 当存在余额列且金额列全为非负值时, 用余额变动方向推断收/支:
    #   cur_bal > prev_bal → 收入;  cur_bal < prev_bal → 支出;  相等 → 不计。
    idx_balance = next((i for i, nm in enumerate(names) if "余额" in nm), None)
    _all_positive = (
        idx_amount is not None and idx_balance is not None
        and all(
            (lambda f: f is None or f >= 0)(_num(r[idx_amount]) if idx_amount < len(r) else None)
            for r in rows
        )
        and any(
            _num(r[idx_amount]) is not None for r in rows if idx_amount < len(r)
        )
    )
    total_d = total_c = 0.0
    in_amount = 0.0
    out_amount = 0.0
    prev_bal = None
    for rec in rows:
        if idx_debit is not None and idx_debit < len(rec):
            f = _num(rec[idx_debit])
            if f is not None:
                total_d += f
        if idx_credit is not None and idx_credit < len(rec):
            f = _num(rec[idx_credit])
            if f is not None:
                total_c += f
        if idx_income is not None and idx_income < len(rec):
            f = _num(rec[idx_income])
            if f is not None:
                in_amount += f
        if idx_expense is not None and idx_expense < len(rec):
            f = _num(rec[idx_expense])
            if f is not None:
                out_amount += f
        if idx_amount is not None and idx_amount < len(rec):
            f = _num(rec[idx_amount])
            if f is not None:
                if _all_positive:
                    bal = _num(rec[idx_balance]) if idx_balance < len(rec) else None
                    if bal is not None and prev_bal is not None:
                        delta = round(bal * 100) - round(prev_bal * 100)
                        if delta > 0:
                            in_amount += f
                        elif delta < 0:
                            out_amount += f
                    prev_bal = bal
                else:
                    if f >= 0:
                        in_amount += f
                    else:
                        out_amount += -f

    report.append(f"数据行数: {n}")
    if idx_debit:
        report.append(f"借方发生额合计: {total_d:,.2f}")
    if idx_credit:
        report.append(f"贷方发生额合计: {total_c:,.2f}")
    if idx_income is not None or idx_expense is not None:
        report.append(f"收入合计: {in_amount:,.2f}  支出合计: {out_amount:,.2f}")
    stats = {
        "rows": n,
        "debit": total_d if idx_debit else None,
        "credit": total_c if idx_credit else None,
        "income": in_amount if (idx_amount is not None or idx_income is not None) else None,
        "expense": out_amount if (idx_amount is not None or idx_expense is not None) else None,
    }
    return report, stats


# ---------- 核心转换(单文件 main 与批量 batch 共用) ----------
def load_columns_template(path):
    """读取视觉锚点 onboarding 生成的列模板 JSON:
    {"columns": [{"name": ..., "x0": ..., "x1": ...}, ...]} → [(lo, hi, name)]。"""
    import json
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    cols = []
    for c in data.get("columns", []):
        cols.append((float(c["x0"]), float(c["x1"]), c["name"]))
    if not cols:
        raise ValueError(f"列模板为空: {path}")
    return cols


def _vision_fallback_extract(doc, patterns, debug=False, password=None,
                             original_error=None):
    """P0-1 表头识别失败自动视觉兜底: 候选页渲染 → 视觉读表头列名 →
    锚点反查列模板 → 带模板重试 extract_statement; 全部失败抛回原错误。

    原则: 视觉只做"每文件一次"的高层判断(读表头列名), 列定位/边界精修/全量
    提取仍由规则管道(ground_header_anchors + build_columns_from_anchors +
    refine_cols_with_data_x)完成。"""
    if PYODIDE:
        # H5 运行时无视觉能力: 直接抛回原错误走纯规则失败路径
        if original_error is not None:
            raise original_error
        raise RuntimeError("视觉兜底未启用(PYODIDE 模式); 请用纯规则路径")
    from vision_utils import ask_header_columns, get_vision_provider  # 延迟导入, 无视觉环境时仍可转
    if get_vision_provider() == "none":
        if original_error is not None:
            raise original_error
        raise RuntimeError("视觉兜底未启用(BANK_PDF_VISION_PROVIDER=none); 请用 --no-vision-fallback 或纯规则路径")
    page_words_list = []
    for pno in range(len(doc)):
        pw = doc[pno].get_text("words")
        if doc[pno].rotation != 0:
            pw = normalize_page_words(doc[pno], pw)
        page_words_list.append(pw)
    max_pages = min(len(doc), 5)
    for pno in range(max_pages):
        anchors = ask_header_columns(doc, pno=pno, retries=1, debug=debug)
        if not anchors:
            log(f"[DEBUG] 视觉兜底: 第{pno+1}页视觉未读出行列名, 试下一页", debug)
            continue
        grounded = ground_anchors_on_page(page_words_list[pno], anchors, debug=debug)
        if grounded is None:
            log(f"[DEBUG] 视觉兜底: 第{pno+1}页锚点反查失败, 试下一页", debug)
            continue
        _hy, _band, cols_raw, _band_bottom = grounded
        cols = build_columns_from_anchors(cols_raw)
        try:
            return extract_statement(
                doc, patterns, debug=debug, columns_template=cols,
                header_anchors=anchors)
        except RuntimeError as e2:
            log(f"[DEBUG] 视觉兜底: 第{pno+1}页模板提取失败: {e2}", debug)
            continue
    if original_error is not None:
        raise original_error
    raise RuntimeError("视觉兜底失败: 未能识别表头, 请人工检查 PDF 是否为对账单")


def classify_doc(doc, debug=False, use_vision=True):
    """文档首页快速分类(P1-4): 表格页 / 封面或条款页 / 扫描件(纯图片) /
    回单或汇总表 / 多账户合并表。

    确定性前置判定(免费且稳定)优先, 视觉只做规则无法区分的判断:
      - 首页无文字层(<5 词) → 扫描件(再让视觉复核, 确认不是封面图);
      - 首页命中表头行 → 表格页;
      - 首页含"微信支付交易明细证明" → 表格页(微信分支);
      - 其余(封面/条款/回单等) → 视觉判断。
    返回 dict: {"type", "reason", "vision_text"}。"""
    try:
        p0_words = doc[0].get_text("words")
    except Exception:  # noqa: BLE001
        p0_words = []
    if doc[0].rotation != 0:
        p0_words = normalize_page_words(doc[0], p0_words)
    text0 = " ".join(w[4] for w in p0_words)
    if WECHAT_SIGN in text0:
        return {"type": "table", "reason": "微信支付交易明细证明", "vision_text": ""}
    if detect_header_line(p0_words) is not None:
        return {"type": "table", "reason": "首页命中表头行", "vision_text": ""}
    if len(p0_words) < 5:
        # 首页文字层几乎为空: 扫描件(纯图片)或极简封面, 均需先确认能否提取。
        # 不做视觉复核——扫描件视觉上也是"表格页", 但无文字层无法规则提取。
        return {"type": "scan", "reason": "首页几乎无文字层(纯图片)", "vision_text": ""}
    if not use_vision or PYODIDE:
        return {"type": "unknown", "reason": "有文字但未命中表头, 无法确定", "vision_text": ""}
    try:
        from vision_utils import call_vision, render_page_png
        png = render_page_png(doc, 0)
        try:
            ans = call_vision(
                png,
                "这是银行对账单 PDF 的第 1 页。请判断该页类型，只回答一个字母和简短理由："
                "A 交易明细表格页(含列名表头和交易数据行)  B 封面或条款说明页  "
                "C 纯扫描图片(照片/影印, 无任何文字)  D 回单/汇总表/多账户合并表。"
                "格式如: B, 这是封面",
                retries=1,
            )
        finally:
            try:
                os.remove(png)
            except OSError:
                pass
        t = (ans or "")[0].upper()
        m = {"A": "table", "B": "cover", "C": "scan", "D": "summary"}
        if t in m:
            return {"type": m[t], "reason": f"视觉判定: {ans}", "vision_text": ans}
        return {"type": "unknown", "reason": f"视觉回答无法解析: {ans}", "vision_text": ans}
    except Exception as e:  # noqa: BLE001
        if debug:
            print(f"[DEBUG] 视觉分类失败: {e}")
        return {"type": "unknown", "reason": f"视觉调用失败: {e}", "vision_text": ""}


def load_descriptor(path_or_dict):
    """加载格式描述符(JSON 路径或已解析 dict), 供 convert_pdf 驱动提取。"""
    if isinstance(path_or_dict, dict):
        return path_or_dict
    with open(path_or_dict, encoding="utf-8") as f:
        return json.load(f)


# ---------- 失败诊断包 + 自动 onboarding 升级(2026-08-20) ----------
# 过程级优化: 失败时自动输出"最小诊断包"(阶段/页号/期望-实际词/建议/下一步命令),
# 并用 onboard_format.py --vision 一次视觉成型生成描述符后自动重试,
# 替代"人工多次试错"(硬格式历史平均 7-18 次尝试)。

DIAG_RULES = [
    ("PDF 密码错误", "password",
     "正确的打开密码", "核对文件名/用户提供的密码后重试(--password <密码>)"),
    ("PDF 已加密", "password",
     "打开密码", "加密对账单(如华夏银行)需 --password; 文件名常含'密码：xxx'"),
    ("需先 OCR", "scan",
     "含文字层的 PDF", "扫描件/纯图片页不支持, 需先 OCR 补文字层后重试"),
    ("未能自动识别表头行", "header",
     "表头行(含各列列名)", "已自动尝试视觉兜底+onboarding; 仍失败用 --header-keywords,"
     "或手动跑 onboard_format.py --vision 生成描述符后 --descriptor 驱动"),
    ("视觉兜底失败", "header",
     "表头行(含各列列名)", "视觉未能读出列名, 检查页面是否被印章大面积遮挡/模糊;"
     "可换更高分辨率截图后用 --anchors 提供列名"),
    ("未能提取到任何", "records",
     "记录行(日期锚点)", "日期格式未命中, 用 --date-pattern 指定"
     "(如 ^\\d{8}$); 或 onboard --vision 生成描述符自动带上日期形态"),
    ("PDF 为空", "empty",
     "非空 PDF", "文件无页面, 检查 PDF 是否损坏"),
]


def classify_failure(err_text):
    """按错误文本归类失败阶段, 返回 (stage, expected, suggestion)。"""
    for sub, stage, expected, suggestion in DIAG_RULES:
        if sub in err_text:
            return stage, expected, suggestion
    return "unknown", "可正常提取的数据", "查看 --debug 输出与诊断包中的页面词样, 定位具体阶段"


def _fp_from_doc(doc, limit=2000):
    """首页文字层指纹(仅用于自动 onboarding 描述符缓存命名)。"""
    try:
        w = doc[0].get_text("words")
        sig = " ".join(x[4] for x in w)[:limit]
        return hashlib.sha256(sig.encode("utf-8")).hexdigest()[:16]
    except Exception:  # noqa: BLE001
        return None


def _desc_columns(desc):
    return [
        (float(c["x0"]), float(c["x1"]), c["name"])
        for c in desc.get("columns", [])
        if "x0" in c and "x1" in c and "name" in c
    ]


def _desc_semantics(desc):
    cols = desc.get("columns", [])
    sems = [c.get("semantic") for c in cols if isinstance(c, dict)]
    if sems and len(sems) == len(cols) and all(sems):
        return sems
    return None


def _desc_anchors(desc):
    """描述符内嵌的视觉表头锚点列表(2026-08-20 起写入; 旧描述符无此字段)。"""
    if isinstance(desc, dict):
        anchors = desc.get("anchors")
        if isinstance(anchors, dict) and isinstance(anchors.get("list"), list):
            return [a for a in anchors["list"] if isinstance(a, str) and a]
    return None


def _auto_onboard_retry(pdf_path, password, cache_dir, doc_fp, debug=False):
    """自动 onboarding(一次视觉成型): 优先 vision 读表头, 失败回退 heuristic。
    成功: 写缓存描述符, 返回 (desc_path, desc, None); 全部失败: (None, None, err)。
    PYODIDE 模式: 无视觉环境且无文件系统, 直接放弃(纯规则路径失败即失败)。"""
    if PYODIDE:
        return None, None, "pyodide: no-vision/no-fs, auto-onboard disabled"
    if doc_fp:
        cache_dir = cache_dir or os.path.join(
            os.path.dirname(os.path.abspath(pdf_path)), "_descriptors")
        desc_path = os.path.join(cache_dir, f"desc_{doc_fp}.json")
        if os.path.exists(desc_path):
            try:
                with open(desc_path, encoding="utf-8") as f:
                    return desc_path, json.load(f), None
            except Exception:  # noqa: BLE001
                pass
    else:
        desc_path = None
    errors = []
    try:
        from vision_utils import get_vision_provider
        _vp = get_vision_provider()
    except Exception:  # noqa: BLE001
        _vp = "visionjs"
    modes = ("heuristic",) if _vp == "none" else ("vision", "heuristic")
    for mode in modes:
        try:
            import onboard_format as ob  # 延迟导入避免循环依赖
            desc = ob.build_format_descriptor(
                pdf_path, password=password, mode=mode, debug=debug)
            if not desc or not desc.get("columns"):
                raise RuntimeError("描述符无列模板")
            if desc_path:
                try:
                    os.makedirs(os.path.dirname(desc_path), exist_ok=True)
                    with open(desc_path, "w", encoding="utf-8") as f:
                        json.dump(desc, f, ensure_ascii=False, indent=2)
                except Exception:  # noqa: BLE001
                    pass
            log(f"[提示] 自动 onboarding({mode})成功, 描述符: {desc_path or '未缓存'}", debug)
            return desc_path, desc, None
        except Exception as e:  # noqa: BLE001
            errors.append(f"{mode}: {e}")
            log(f"[提示] 自动 onboarding({mode})失败: {e}", debug)
    return None, None, "; ".join(errors)


def _extract_general_with_escalation(doc, pdf_path, patterns, *,
                                     debug=False, columns_template=None,
                                     layout=None, footer_extra=None,
                                     semantics=None, vision_fallback=True,
                                     auto_onboard=True, password=None,
                                     onboard_cache=None, doc_fp=None,
                                     escalation=None, header_anchors=None):
    """通用分支提取 + 两级自动升级:
      1) 表头识别失败 → 视觉兜底(_vision_fallback_extract, 读列名反查列模板);
      2) 仍失败或其它提取失败 → 自动 onboarding(视觉优先, 失败回退启发式),
         生成/缓存格式描述符后用描述符重试(一次视觉成型, 替代多次试错)。"""
    def retry_with_desc(desc, desc_path):
        if escalation is not None:
            escalation.update({"attempted": True, "descriptor": desc_path})
        try:
            retry_patterns = ([re.compile(desc["date_pattern"])]
                              if desc.get("date_pattern") else patterns)
            res = extract_statement(
                doc, retry_patterns, debug=debug,
                columns_template=_desc_columns(desc),
                layout=desc.get("layout"), footer_extra=desc.get("footer_keywords"),
                semantics=_desc_semantics(desc),
                header_anchors=_desc_anchors(desc),
            )
            if escalation is not None:
                escalation["ok"] = True
            return res[0], res[1], res[2], {"onboarded": desc_path}
        except RuntimeError as e3:
            if escalation is not None:
                escalation["error"] = str(e3)
            raise

    try:
        res = extract_statement(
            doc, patterns, debug=debug, columns_template=columns_template,
            layout=layout, footer_extra=footer_extra, semantics=semantics,
            header_anchors=header_anchors)
        return res[0], res[1], res[2], {}
    except RuntimeError as e:
        if vision_fallback and "未能自动识别表头行" in str(e):
            log("[提示] 表头自动识别失败, 启用视觉兜底...", debug)
            try:
                res = _vision_fallback_extract(
                    doc, patterns, debug=debug, password=password,
                    original_error=e)
                return res[0], res[1], res[2], {"fallback": "vision"}
            except RuntimeError as e2:
                e = e2
        if auto_onboard and columns_template is None and layout is None:
            desc_path, desc, onb_err = _auto_onboard_retry(
                pdf_path, password, onboard_cache, doc_fp, debug=debug)
            if desc is not None:
                return retry_with_desc(desc, desc_path)
            if escalation is not None:
                escalation.update(
                    {"attempted": True, "error": onb_err or str(e)})
        raise e


def _diag_actual_samples(pdf_path, stage, password=None, max_pages=3):
    """失败现场取样: 打开 PDF(尽力而为), 记录每页词数/表头命中/日期行数/
    列模板, 供"期望 vs 实际"比对。"""
    samples = {"pages": {}, "note": ""}
    try:
        doc = open_pdf(pdf_path, password=password)
        try:
            if doc.needs_pass:
                if not password or not doc.authenticate(password):
                    samples["note"] = "加密 PDF 未认证, 无法取样; 认证后重跑即可"
                    return samples
            n = len(doc)
            for pno in range(min(n, max_pages)):
                pw = doc[pno].get_text("words")
                if doc[pno].rotation != 0:
                    pw = normalize_page_words(doc[pno], pw)
                page_samp = {"n_words": len(pw)}
                if stage == "header":
                    hdr = detect_header_line(pw)
                    page_samp["header_detected"] = hdr is not None
                    top = sorted(pw, key=lambda x: x[1])[:40]
                    page_samp["top_words"] = [w[4] for w in top[:25]]
                elif stage == "records":
                    date_rows = find_date_rows(pw, DEFAULT_DATE_PATTERNS)
                    page_samp["date_rows_found"] = len(date_rows)
                    hdr = detect_header_line(pw)
                    if hdr is not None:
                        cols, _bb = detect_column_boundaries(pw, hdr)
                        page_samp["cols_detected"] = [c[2] for c in cols][:20]
                    page_samp["date_like_words"] = [
                        w[4] for w in pw if re.match(r"^(19|20)\d{2}", w[4])][:12]
                else:
                    page_samp["first_words"] = [
                        w[4] for w in sorted(pw, key=lambda x: x[1])[:20]]
                samples["pages"][str(pno + 1)] = page_samp
        finally:
            doc.close()
    except Exception as e:  # noqa: BLE001
        samples["note"] = f"取样失败: {e}"
    return samples


def _suggest_command(pdf_path, stage):
    base = os.path.dirname(os.path.abspath(__file__))
    if stage == "header":
        return (f"python {os.path.join(base, 'onboard_format.py')} "
                f"--input \"{pdf_path}\" --vision")
    if stage == "records":
        return (f"python {os.path.abspath(__file__)} --input \"{pdf_path}\" "
                f"--date-pattern \"^\\d{{8}}$\" --debug")
    if stage == "password":
        return (f"python {os.path.abspath(__file__)} --input \"{pdf_path}\" "
                f"--password <密码>")
    return (f"python {os.path.abspath(__file__)} --input \"{pdf_path}\" --debug")


def write_failure_diag(pdf_path, out_path, exc, escalation=None,
                       password=None, debug=False):
    """失败时写 <输出>.diag.json(阶段/页号/期望-实际词/建议/重试命令),
    并渲染失败页 PNG 供人工/视觉复核。写失败不掩盖原始错误。"""
    from datetime import datetime
    err_text = str(exc)
    stage, expected, suggestion = classify_failure(err_text)
    samples = _diag_actual_samples(pdf_path, stage, password=password)
    diag = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "pdf": os.path.abspath(pdf_path),
        "error": err_text,
        "stage": stage,
        "expected": expected,
        "actual": samples,
        "date_patterns_tried": [p.pattern for p in DEFAULT_DATE_PATTERNS],
        "suggestion": suggestion,
        "escalation": escalation or {},
        "retry_command": _suggest_command(pdf_path, stage),
    }
    diag_path = os.path.splitext(out_path)[0] + ".diag.json"
    with open(diag_path, "w", encoding="utf-8") as f:
        json.dump(diag, f, ensure_ascii=False, indent=2)
    failed = samples.get("pages", {})
    page = None
    for pno, ps in failed.items():
        if stage == "header" and ps.get("header_detected") is False:
            page = int(pno)
            break
        if stage == "records" and ps.get("date_rows_found") == 0:
            page = int(pno)
            break
    if page is None and samples.get("pages"):
        page = 1
    if page:
        try:
            doc = open_pdf(pdf_path, password=password)
            try:
                if doc.needs_pass and password:
                    doc.authenticate(password)
                if PYODIDE:
                    # H5 模式: 无文件系统, 跳过失败页 PNG 落盘(前端不依赖它)
                    diag["png"] = None
                else:
                    png = os.path.splitext(out_path)[0] + f"_page{page}.png"
                    doc[page - 1].get_pixmap(
                        matrix=pymupdf.Matrix(1.5, 1.5)).save(png)
                    diag["png"] = png
                    with open(diag_path, "w", encoding="utf-8") as f:
                        json.dump(diag, f, ensure_ascii=False, indent=2)
            finally:
                doc.close()
        except Exception:  # noqa: BLE001
            pass
    log(f"[诊断] 失败阶段: {stage} | 期望: {expected} | 诊断包: {diag_path}", False)
    return diag_path


# ---------- 输出单元格类型化(2026-08-20 常驻) ----------
# 日期/时间列与金额/余额列默认输出为 Excel 真类型(可筛选/排序/求和)。
# 需要与旧版本一致的纯文本输出时, 用 --keep-text。
DATE_COL_HINTS = ("日期", "时间")
MONEY_COL_HINTS = ("金额", "余额", "发生额", "收入", "支出", "借方", "贷方", "手续费")
_DATE_PARSE_PATTERNS = (
    ("%Y-%m-%d %H:%M:%S", True),
    ("%Y-%m-%d %H:%M", True),
    ("%Y/%m/%d %H:%M:%S", True),
    ("%Y/%m/%d %H:%M", True),
    ("%Y%m%d %H:%M:%S", True),
    ("%Y%m%d %H:%M", True),
    ("%Y-%m-%d", False),
    ("%Y/%m/%d", False),
    ("%Y%m%d", False),
    ("%Y.%m.%d", False),
)


def _to_date_value(v):
    """可解析的日期/时间文本 → date/datetime; 无法解析时原样返回。"""
    if v is None or isinstance(v, (date, datetime)):
        return v
    s = str(v).strip()
    if not s:
        return v
    # 农业银行交易时间列因 PDF 日期跨行换行产生 "2025-01- 02": 去掉分隔符后空格
    s = re.sub(r"(?<=\d)([-/.])\s+(?=\d)", r"\1", s)
    # 6 位纯数字(如浦发 '093351' = 09:33:51)是时间不是日期: 若直接进 %Y%m%d
    # strptime, 正则回溯会解析成 '0933-05-01' 这类荒谬日期。先按时间词拦截,
    # 合法 6 位时分秒保持原文本(时间列语义)。
    if re.fullmatch(r"\d{6}", s) and is_time_word(s):
        return s
    for fmt, has_time in _DATE_PARSE_PATTERNS:
        try:
            dt = datetime.strptime(s, fmt)
        except ValueError:
            continue
        return dt if has_time else dt.date()
    m = re.match(
        r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})"
        r"(?:[ T](\d{1,2}):(\d{2})(?::(\d{2}))?)?$", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            if m.group(4):
                return datetime(y, mo, d, int(m.group(4)), int(m.group(5)),
                                int(m.group(6) or 0))
            return date(y, mo, d)
        except ValueError:
            return v
    return v


def _to_money_value(v):
    """金额/余额文本 → float(剥离千分位/货币符/前缀, 保留正负号); 无法解析时原样返回。"""
    if v is None or isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if not s:
        return v
    t = re.sub(r"[\s,，￥¥]", "", s)
    # 2026-08-25: 兼容 PDF 无小数位金额词(建行企业明细 '50000.' 等, 源 PDF
    # 渲染省略 .00): (?:\.\d+)? → (?:\.\d*)?, 允许小数点后 0 位(仍整词锚定)。
    m = re.search(r"[-+]?\d+(?:\.\d*)?$", t)
    if not m:
        return v
    try:
        return float(m.group(0))
    except ValueError:
        return v


def typed_records(header_names, records, keep_text=False):
    """按表头语义把日期/时间列与金额/余额列转成 date/datetime/float。

    默认开启(2026-08-20 常驻); keep_text=True 时保留 PDF 原文文本, 与旧版一致。
    """
    if keep_text or not records:
        return records
    flags = [
        ("D" if any(k in str(name) for k in DATE_COL_HINTS)
         else "M" if any(k in str(name) for k in MONEY_COL_HINTS) else None)
        for name in header_names
    ]
    out = []
    for rec in records:
        r = list(rec)
        for i, flag in enumerate(flags):
            if i >= len(r):
                break
            if flag == "D":
                r[i] = _to_date_value(r[i])
            elif flag == "M":
                r[i] = _to_money_value(r[i])
        out.append(tuple(r))
    return out


def _apply_cell_formats(ws, header_names, n_data):
    """为日期/时间列与金额/余额列设置 Excel 数字格式(与 typed_records 配套)。"""
    fmt_by_col = []
    for name in header_names:
        if any(k in str(name) for k in DATE_COL_HINTS):
            fmt_by_col.append("DATE")
        elif any(k in str(name) for k in MONEY_COL_HINTS):
            fmt_by_col.append("MONEY")
        else:
            fmt_by_col.append(None)
    for row in ws.iter_rows(min_row=2, max_row=n_data):
        for i, cell in enumerate(row):
            if i >= len(fmt_by_col) or fmt_by_col[i] is None:
                continue
            v = cell.value
            if isinstance(v, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif isinstance(v, date):
                cell.number_format = "yyyy-mm-dd"
            elif fmt_by_col[i] == "MONEY" and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00;[Red]-#,##0.00"


def _output_layout(header_names):
    """列对齐与列宽(openpyxl 与 xlsxwriter 两个写出后端共用)。
    返回 (aligns, widths): aligns[i] = (horizontal|None, wrap)。"""
    aligns = []
    widths = []
    for name in header_names:
        if any(k in name for k in ("摘要", "户名", "行名", "账号", "用途", "备注")):
            aligns.append((None, True))
            widths.append(32)
        elif any(k in name for k in ("发生额", "金额", "余额")):
            aligns.append(("right", False))
            widths.append(14)
        elif "时间" in name:
            aligns.append(("center", False))
            widths.append(20)
        else:
            aligns.append(("center", False))
            widths.append(14)
    return aligns, widths


def _write_xlsx_streaming(out_path, sheet_name, header_names, typed,
                          aligns, widths, keep_text):
    """xlsxwriter 流式写出(constant_memory 逐行落盘, 2026-08-21)。
    输出契约与 openpyxl 路径一致: 表头蓝底白字加粗居中、数据行细边框 + 按列对齐、
    日期/金额列 Excel 数字格式(--keep-text 时不设)、列宽、冻结首行。
    数字格式按单元格值类型选择(日期列可同时含 date 与 datetime 行)。"""
    import xlsxwriter
    wb = xlsxwriter.Workbook(out_path, {"constant_memory": True})
    ws = wb.add_worksheet(sheet_name[:31])
    hdr = wb.add_format({
        "bold": True, "font_color": "#FFFFFF", "bg_color": "#305496",
        "align": "center", "valign": "vcenter", "text_wrap": True,
        "border": 1, "border_color": "#BFBFBF",
    })
    fmt_kind = []
    for name in header_names:
        if any(k in str(name) for k in DATE_COL_HINTS):
            fmt_kind.append("D")
        elif any(k in str(name) for k in MONEY_COL_HINTS):
            fmt_kind.append("M")
        else:
            fmt_kind.append(None)
    base_fmts = {}
    date_fmts = {}
    dtime_fmts = {}
    money_fmts = {}
    for ci, (name, (h, wrap)) in enumerate(zip(header_names, aligns)):
        kw = {"border": 1, "border_color": "#BFBFBF", "valign": "vcenter"}
        if h:
            kw["align"] = h
        if wrap:
            kw["text_wrap"] = True
        base_fmts[ci] = wb.add_format(kw)
        kind = fmt_kind[ci]
        if kind == "D":
            date_fmts[ci] = wb.add_format(dict(kw, num_format="yyyy-mm-dd"))
            dtime_fmts[ci] = wb.add_format(
                dict(kw, num_format="yyyy-mm-dd hh:mm:ss"))
        elif kind == "M":
            money_fmts[ci] = wb.add_format(
                dict(kw, num_format="#,##0.00;[Red]-#,##0.00"))
    ws.write_row(0, 0, header_names, hdr)
    for r, rec in enumerate(typed, start=1):
        for c, v in enumerate(rec):
            f = base_fmts.get(c, base_fmts[0] if base_fmts else None)
            if not keep_text:
                kind = fmt_kind[c] if c < len(fmt_kind) else None
                if kind == "D":
                    if isinstance(v, datetime):
                        f = dtime_fmts.get(c, f)
                    elif isinstance(v, date):
                        f = date_fmts.get(c, f)
                elif kind == "M" and isinstance(v, (int, float)):
                    f = money_fmts.get(c, f)
            ws.write(r, c, v, f)
    for i, w in enumerate(widths):
        ws.set_column(i, i, w)
    ws.freeze_panes(1, 0)  # 冻结首行(= openpyxl freeze_panes="A2")
    wb.close()


def _write_xlsx_openpyxl(out_path, sheet_name, header_names, typed,
                         aligns, widths, keep_text):
    """openpyxl 写出回退路径(环境无 xlsxwriter 时)。与原实现行为一致:
    表头样式 + 数据行 StyleArray 快速路径(borderId/alignmentId 直设) +
    日期/金额数字格式 + 列宽 + 冻结首行。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(header_names)
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="305496")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for rec in typed:
        ws.append(rec)
    align_by_col = []
    for h, wrap in aligns:
        if h:
            align_by_col.append(Alignment(horizontal=h, vertical="center"))
        else:
            align_by_col.append(Alignment(vertical="center", wrap_text=wrap))
    n_data = ws.max_row
    if n_data >= 2:  # records 非空时恒成立, 防御性判断
        first_cells = ws[2]
        for cell in first_cells:
            cell.border = border
            ci = cell.column - 1
            cell.alignment = (align_by_col[ci] if ci < len(align_by_col)
                              else Alignment(horizontal="center", vertical="center"))
        row1_ids = [(c._style.borderId, c._style.alignmentId) for c in first_cells]
        for row in ws.iter_rows(min_row=3, max_row=n_data):
            for i, cell in enumerate(row):
                if cell._style is None:
                    cell._style = StyleArray()
                if i < len(row1_ids):
                    cell._style.borderId = row1_ids[i][0]
                    cell._style.alignmentId = row1_ids[i][1]
                else:
                    cell._style.borderId = row1_ids[-1][0]
                    cell._style.alignmentId = row1_ids[-1][1]
        if not keep_text:
            _apply_cell_formats(ws, header_names, n_data)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(out_path)


def _write_xlsx(out_path, sheet_name, header_names, typed, aligns, widths, keep_text):
    """写 xlsx: 优先 xlsxwriter 流式(constant_memory, 大文件快 0.5-1s);
    环境缺 xlsxwriter 时回退 openpyxl 路径(输出契约一致)。"""
    try:
        import xlsxwriter  # noqa: F401
    except Exception:  # noqa: BLE001
        xlsxwriter = None
    if xlsxwriter is not None:
        _write_xlsx_streaming(out_path, sheet_name, header_names, typed,
                              aligns, widths, keep_text)
    else:
        _write_xlsx_openpyxl(out_path, sheet_name, header_names, typed,
                             aligns, widths, keep_text)


def _convert_pdf_core(pdf_path, out_path=None, sheet=None, date_pattern=None,
                      header_keywords=None, debug=False, password=None,
                      columns_template=None, descriptor=None, vision_fallback=True,
                      quick_classify=False, wechat_anchors=None, auto_onboard=True,
                      diag=True, onboard_cache=None, keep_text=False,
                      write_meta=False, _escalation=None):
    """转换核心: 打开 PDF(支持加密) → 分流(微信/通用) → 提取 → 写 xlsx → 统计。
    返回 (report, stats, meta, out_path, sheet_name, n_cols)。
    batch_convert.py 直接复用本函数, 避免 subprocess 双层进程开销。

    新增参数(2026-08-19 升级):
      descriptor     格式描述符 dict/path: 提供 columns/date_pattern/layout/
                     footer_keywords/semantics, 提取跳过对应启发式试探;
      vision_fallback  表头识别失败时自动视觉兜底(默认开, --no-vision-fallback 关);
      quick_classify  提取前对首页做文档类型快速分类(扫描件直接报"需先 OCR")。

    新增参数(2026-08-20 升级, 公开入口 convert_pdf 透传):
      auto_onboard   启发式+视觉兜底仍失败时, 自动 onboarding(视觉优先/启发式回退)
                     生成格式描述符并缓存, 用描述符重试(默认开, --no-auto-onboard 关);
      diag           失败时自动写 <输出>.diag.json 诊断包(默认开, --no-diag 关);
      onboard_cache  自动 onboarding 描述符缓存目录(默认 <PDF目录>/_descriptors)。"""
    if not PYODIDE:
        pdf_path = os.path.abspath(pdf_path)
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"文件不存在: {pdf_path}")

    # 格式描述符(优先于 CLI 单参数; CLI 单参数显式给出时以 CLI 为准)
    desc = load_descriptor(descriptor) if descriptor else None
    if desc:
        if date_pattern is None and desc.get("date_pattern"):
            date_pattern = desc["date_pattern"]
        if columns_template is None and desc.get("columns"):
            columns_template = [
                (float(c["x0"]), float(c["x1"]), c["name"])
                for c in desc["columns"] if "x0" in c and "x1" in c and "name" in c
            ]
    desc_layout = desc.get("layout") if desc else None
    desc_footer = desc.get("footer_keywords") if desc else None
    desc_semantics = (
        [c.get("semantic") for c in desc.get("columns", []) if isinstance(c, dict)]
        if desc and desc.get("columns") else None
    )
    if desc_semantics and any(s is None for s in desc_semantics):
        desc_semantics = None
    if columns_template is not None and desc_semantics is not None \
            and len(desc_semantics) != len(columns_template):
        desc_semantics = None

    # 日期模式
    patterns = [re.compile(date_pattern)] if date_pattern else DEFAULT_DATE_PATTERNS
    # 表头关键词
    if header_keywords:
        global DEFAULT_HEADER_KEYWORDS
        DEFAULT_HEADER_KEYWORDS = [k.strip() for k in header_keywords.split(",")]

    # 输出路径(PYODIDE 模式 out_path 为 io.BytesIO; 本地模式为路径)
    if not PYODIDE:
        out = os.path.abspath(out_path) if out_path else os.path.splitext(pdf_path)[0] + ".xlsx"
    else:
        out = out_path

    # 打开一次 PDF(加密文件用 --password 认证), 分流微信/通用
    doc = open_pdf(pdf_path, password=password)
    try:
        if doc.needs_pass:
            if not password:
                raise RuntimeError("PDF 已加密, 请用 --password 提供打开密码")
            if not doc.authenticate(password):
                raise RuntimeError("PDF 密码错误")
        if quick_classify:
            try:
                if PYODIDE:
                    _quick_use_vision = False
                else:
                    from vision_utils import get_vision_provider
                    _quick_use_vision = get_vision_provider() != "none"
            except Exception:  # noqa: BLE001
                _quick_use_vision = not PYODIDE
            cls = classify_doc(doc, debug=debug, use_vision=_quick_use_vision)
            ctype = cls.get("type") if isinstance(cls, dict) else None
            if ctype == "scan":
                raise RuntimeError(
                    "检测到扫描件/纯图片页(无文字层), 需先 OCR 补充文字层后重试")
            if ctype in ("cover", "summary"):
                log(f"[提示] 首页为{('封面/条款页' if ctype == 'cover' else '回单/汇总表')}, "
                    f"将从含表头页开始提取: {cls.get('reason', '')}", debug)
        if is_wechat_doc(doc):
            wc_bounds = None
            wc_anchors = None
            if desc and desc.get("layout") == "vertical" and desc.get("columns"):
                wc_bounds = [
                    (float(c["x0"]), float(c["x1"]), c["name"])
                    for c in desc["columns"]
                    if "x0" in c and "x1" in c and "name" in c
                ]
                if len(wc_bounds) != len(WECHAT_COLS):
                    wc_bounds = None
            if wc_bounds is None and wechat_anchors:
                wc_anchors = [a.strip() for a in wechat_anchors.split(",") if a.strip()]
            header_names, records, meta = extract_wechat_statement(
                doc, debug=debug, anchors=wc_anchors, bounds=wc_bounds)
            is_wechat = True
        else:
            header_names, records, meta, _esc_extra = _extract_general_with_escalation(
                doc, pdf_path, patterns, debug=debug,
                columns_template=columns_template, layout=desc_layout,
                footer_extra=desc_footer, semantics=desc_semantics,
                vision_fallback=vision_fallback, auto_onboard=auto_onboard,
                password=password, onboard_cache=onboard_cache,
                doc_fp=_fp_from_doc(doc) if auto_onboard else None,
                escalation=_escalation,
                header_anchors=_desc_anchors(desc) if desc else None)
            if _esc_extra.get("onboarded"):
                log(f"[提示] 自动 onboarding 成功并已用描述符重试: "
                    f"{_esc_extra['onboarded']}", debug)
            is_wechat = False
    finally:
        doc.close()

    # 写 xlsx(2026-08-21: xlsxwriter 流式 constant_memory 逐行落盘, 大文件快
    # 0.5-1s; 环境无 xlsxwriter 时回退 openpyxl 路径, 输出契约一致)
    if PYODIDE:
        # H5 模式: pdf_path 是 str(pdf_bytes) 哨兵不可用于取名, 由 shim 传 sheet
        sheet_name = (sheet or "对账单")[:31]
    else:
        sheet_name = (sheet or os.path.splitext(os.path.basename(pdf_path))[0])[:31]
    typed = typed_records(header_names, records, keep_text=keep_text)
    aligns, widths = _output_layout(header_names)
    _write_xlsx(out, sheet_name, header_names, typed, aligns, widths, keep_text)

    # 提取统计(直接基于记录行, 与写出后端解耦, 不再依赖内存工作表)
    if is_wechat:
        report, stats = compute_wechat_stats(header_names, typed)
    else:
        report, stats = compute_stats(header_names, typed, semantics=desc_semantics)
    # 页面级元数据 sidecar(2026-08-21): 已知格式快速核对(qa_vision --fast)直接
    # 复用每页记录数/首末日期/首末流水号/首末余额, 免去二次转换与全页重扫描。
    if write_meta and not PYODIDE:
        try:
            meta_path = os.path.splitext(out)[0] + ".meta.json"
            with open(meta_path, "w", encoding="utf-8") as f:
                json.dump({
                    "wechat": is_wechat,
                    "page_count": meta.get("page_count"),
                    "records": stats.get("rows"),
                    "report": report,
                    "pages": meta.get("pages", []),
                }, f, ensure_ascii=False, indent=2)
        except Exception:  # noqa: BLE001 (元数据写失败不掩盖转换结果)
            pass
    return report, stats, meta, out, sheet_name, len(header_names)


def convert_pdf(pdf_path, out_path=None, sheet=None, date_pattern=None,
                header_keywords=None, debug=False, password=None,
                columns_template=None, descriptor=None, vision_fallback=True,
                quick_classify=False, wechat_anchors=None, auto_onboard=True,
                diag=True, onboard_cache=None, keep_text=False, write_meta=False,
                vision_provider="auto"):
    """公开入口(2026-08-20 起): 完整转换流程 + 失败自动诊断包。

    调用 _convert_pdf_core 完成提取/写盘; 任何失败(密码/扫描件/表头/日期/记录)
    自动写 <输出>.diag.json(阶段/页号/期望-实际词/建议/重试命令) + 失败页 PNG,
    再抛出原始错误。失败路径默认开启 auto_onboard(视觉 onboarding 一次成型→
    描述符重试), 正常路径零额外开销。

    新增参数:
      auto_onboard   启发式+视觉兜底仍失败时自动 onboarding 并描述符重试(默认开);
      diag           失败时写诊断包(默认开, --no-diag 关闭);
      onboard_cache  描述符缓存目录(默认 <PDF目录>/_descriptors);
      vision_provider "visionjs"(默认, 调外部 vision.js) 或
                      "model"(原生多模态模型读图, 见 SKILL.md 视觉角色节)。
                      model 下视觉兜底/onboarding 不调外部视觉, 而是写待读图请求
                      并抛 ModelVisionPendingError, 待代理读图写回答案后重跑命中。"""
    if not PYODIDE:
        out = os.path.abspath(out_path) if out_path else \
            os.path.splitext(os.path.abspath(pdf_path))[0] + ".xlsx"
    else:
        out = out_path  # PYODIDE: out_path 即 io.BytesIO
    escalation = {"attempted": False, "ok": False, "descriptor": None, "error": None}
    # 统一解析 provider: auto 在外部 vision.js 不可用时自动降级 none,
    # 无图像能力环境不再默认触发外部视觉或进入“模型读图待办”死循环。
    resolved_provider = "none"
    if not PYODIDE:
        try:
            from vision_utils import set_vision_provider
            resolved_provider = set_vision_provider(vision_provider)
        except Exception:  # noqa: BLE001
            resolved_provider = "none"
            os.environ["BANK_PDF_VISION_PROVIDER"] = "none"
    if resolved_provider == "none":
        vision_fallback = False
        log("[提示] 视觉功能不可用或已禁用, 使用纯规则/文字层路径; "
            "表头失败时禁用自动视觉兜底", debug)
    try:
        return _convert_pdf_core(
            pdf_path, out_path=out_path, sheet=sheet, date_pattern=date_pattern,
            header_keywords=header_keywords, debug=debug, password=password,
            columns_template=columns_template, descriptor=descriptor,
            vision_fallback=vision_fallback, quick_classify=quick_classify,
            wechat_anchors=wechat_anchors, auto_onboard=auto_onboard,
            diag=False, onboard_cache=onboard_cache, keep_text=keep_text,
            write_meta=write_meta, _escalation=escalation)
    except Exception as e:
        # 原生多模态模型回环待办不是失败: 不写诊断包, 交给 main 转退出码 3。
        if "ModelVisionPendingError" in type(e).__name__:
            raise
        if diag and not PYODIDE:
            try:
                write_failure_diag(pdf_path, out, e, escalation=escalation,
                                   password=password, debug=debug)
            except Exception as de:  # noqa: BLE001
                log(f"[提示] 诊断包写入失败: {de}", debug)
        raise


# ---------- 主流程 ----------
def main():
    import sys
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    ap = argparse.ArgumentParser(description="银行对账单 PDF → Excel")
    ap.add_argument("--input", help="输入 PDF")
    ap.add_argument("--output", help="输出 xlsx(默认同 PDF 目录)")
    ap.add_argument("--date-pattern", help="日期行匹配正则")
    ap.add_argument("--header-keywords", help="表头关键词, 逗号分隔")
    ap.add_argument("--columns-json", help="视觉锚点 onboarding 生成的列模板 JSON")
    ap.add_argument("--descriptor", help="格式描述符 JSON(onboarding 生成, 含 date_pattern/layout/columns/semantic)")
    ap.add_argument("--password", help="PDF 打开密码(加密对账单)")
    ap.add_argument("--sheet", help="工作表名")
    ap.add_argument("--no-vision-fallback", action="store_true",
                    help="关闭表头识别失败的自动视觉兜底")
    ap.add_argument("--no-auto-onboard", action="store_true",
                    help="关闭失败时自动 onboarding(视觉一次成型→描述符重试)")
    ap.add_argument("--no-diag", action="store_true",
                    help="关闭失败诊断包(<输出>.diag.json + 失败页 PNG)")
    ap.add_argument("--keep-text", action="store_true",
                    help="日期/金额列保留 PDF 原文文本(默认输出为 Excel 日期/数值格式)")
    ap.add_argument("--onboard-cache", help="自动 onboarding 描述符缓存目录"
                    "(默认 <PDF目录>/_descriptors)")
    ap.add_argument("--write-meta", action="store_true",
                    help="转换时顺带写 <输出>.meta.json(页面级元数据, 供 qa_vision --fast 复用)")
    ap.add_argument("--quick-classify", action="store_true",
                    help="提取前对首页做文档类型快速分类(扫描件直接报需 OCR)")
    ap.add_argument("--wechat-anchors", help="微信表头列名锚点(逗号分隔), 动态列锚点, 失败回退硬编码")
    ap.add_argument("--no-vision", dest="vision_provider",
                    action="store_const", const="none",
                    help="关闭一切视觉能力, 使用纯规则/文字层路径(等价 --vision-provider none)")
    ap.add_argument("--vision-provider", default="auto",
                    choices=["auto", "visionjs", "model", "none"],
                    help="视觉 provider: auto(外部 vision.js 可用则用, 否则自动关闭); "
                         "visionjs(默认外部 vision.js); model(原生多模态模型读图, "
                         "显式写待读图请求后退出码 3); none(无图像能力安全降级, "
                         "只走确定性文字层路径)")
    ap.add_argument("--debug", action="store_true")
    args = ap.parse_args()

    if not args.input:
        print("错误: 缺少 --input", file=sys.stderr)
        ap.print_usage(file=sys.stderr)
        sys.exit(2)

    try:
        columns_template = load_columns_template(args.columns_json) if args.columns_json else None
        report, stats, meta, out_path, sheet_name, n_cols = convert_pdf(
            args.input,
            out_path=args.output,
            sheet=args.sheet,
            date_pattern=args.date_pattern,
            header_keywords=args.header_keywords,
            debug=args.debug,
            password=args.password,
            columns_template=columns_template,
            descriptor=args.descriptor,
            vision_fallback=not args.no_vision_fallback,
            quick_classify=args.quick_classify,
            wechat_anchors=args.wechat_anchors,
            auto_onboard=not args.no_auto_onboard,
            diag=not args.no_diag,
            onboard_cache=args.onboard_cache,
            keep_text=args.keep_text,
            write_meta=args.write_meta,
            vision_provider=args.vision_provider,
        )
    except Exception as e:
        if "ModelVisionPendingError" in type(e).__name__:
            png = getattr(e, "png", "")
            answer_file = getattr(e, "answer_file", "")
            request_dir = getattr(e, "request_dir", "")
            print(">>> 需要当前模型读图(visual provider=model, 未调外部 vision.js):",
                  file=sys.stderr)
            print(f"    读图: {png}", file=sys.stderr)
            print(f"    问题: {getattr(e, 'prompt', '')}", file=sys.stderr)
            print(f"    请求目录: {request_dir}", file=sys.stderr)
            print(f"    请用模型读图后把答案写入: {answer_file}", file=sys.stderr)
            print(f"    写入后重跑同一命令(含 --vision-provider model)即继续。",
                  file=sys.stderr)
            sys.exit(3)
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(2)

    print(f"输出: {out_path}")
    print(f"工作表: {sheet_name} | 行数: {stats['rows'] + 1} (含表头) | 列数: {n_cols}")
    if meta.get("truncated_words"):
        print(f"PDF源截断词(含省略号, 已剥离%机构代码尾巴): {meta['truncated_words']} 处")
    print("提取统计(交付前请人工与 PDF 合计比对并抽样核对拼接处):")
    for line in report:
        print(f"  {line}")
    sys.exit(0)


if __name__ == "__main__":
    main()
